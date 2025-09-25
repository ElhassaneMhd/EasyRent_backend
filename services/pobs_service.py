"""
POBS Service - Converted from EasyRent_AggiornaPOBS.py
Removes Tkinter GUI and converts to web service functions
"""

import pandas as pd
import os
import shutil
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import numbers
from .logger_service import log_pobs_operation
from .realtime_logger import realtime_logger

def filter_resolved_rejected_status(df, log_function=None):
    """
    Filter out rows with Resolved-Rejected status in both Italian and English
    Column names checked: STATO, STATUS, Stato, Status
    Values filtered: Risolto-Rifiutato, Resolved-Rejected (case insensitive)
    """
    if df.empty:
        return df, 0

    original_count = len(df)

    # Find status columns (both Italian and English)
    status_columns = []
    for col in df.columns:
        if col.upper() in ['STATO', 'STATUS']:
            status_columns.append(col)

    if not status_columns:
        if log_function:
            log_function("[INFO] No STATUS/STATO column found - no filtering applied")
        return df, 0

    # Filter out resolved-rejected records for each status column found
    filtered_df = df.copy()
    excluded_values = ['risolto-rifiutato', 'resolved-rejected']

    for status_col in status_columns:
        # Create case-insensitive mask for filtering
        mask = ~filtered_df[status_col].astype(str).str.strip().str.lower().isin(excluded_values)
        filtered_df = filtered_df[mask]

        if log_function:
            log_function(f"[INFO] Applied status filter on column '{status_col}'")

    filtered_count = len(filtered_df)
    excluded_count = original_count - filtered_count

    if log_function and excluded_count > 0:
        log_function(f"[INFO] Excluded {excluded_count} records with 'Resolved-Rejected' / 'Risolto-Rifiutato' status")

    return filtered_df, excluded_count

def verify_new_records(noleggio_path, pobs_path):
    """
    Verify new records between Noleggio and POBS files
    Converted from verifica_nuovi function
    """
    processing_log = []

    try:
        processing_log.append("[INFO] Starting POBS verification process...")
        chiave = "POBS ID"

        # Read files
        processing_log.append(f"[INFO] Reading Noleggio file: {os.path.basename(noleggio_path)}")
        df_noleggio = pd.read_excel(noleggio_path, dtype=str)
        processing_log.append(f"[OK] Loaded {len(df_noleggio)} records from Noleggio file")

        # Apply status filtering for POBS verification
        processing_log.append("[INFO] Applying status filter to exclude 'Resolved-Rejected' / 'Risolto-Rifiutato' records")
        df_noleggio, excluded_count = filter_resolved_rejected_status(df_noleggio, lambda msg: processing_log.append(msg))
        if excluded_count > 0:
            processing_log.append(f"[INFO] After filtering: {len(df_noleggio)} records remaining (excluded {excluded_count} resolved-rejected records)")
        else:
            processing_log.append("[INFO] No records excluded by status filter")

        processing_log.append(f"[INFO] Reading POBS file: {os.path.basename(pobs_path)}")
        df_pobs = pd.read_excel(pobs_path, dtype=str)
        processing_log.append(f"[OK] Loaded {len(df_pobs)} records from POBS file")


        # Check if required column exists
        processing_log.append(f"[INFO] Checking for required column: '{chiave}'")
        if chiave not in df_noleggio.columns:
            processing_log.append(f"[ERROR] Column '{chiave}' not found in Noleggio file")
            return {
                'success': False,
                'error': f"Column '{chiave}' not found in Noleggio file. Available columns: {list(df_noleggio.columns)}",
                'processing_log': processing_log
            }

        if chiave not in df_pobs.columns:
            processing_log.append(f"[ERROR] Column '{chiave}' not found in POBS file")
            return {
                'success': False,
                'error': f"Column '{chiave}' not found in POBS file. Available columns: {list(df_pobs.columns)}",
                'processing_log': processing_log
            }

        processing_log.append(f"[OK] Required column '{chiave}' found in both files")

        # Clean and normalize data
        processing_log.append("[INFO] Cleaning and normalizing data...")
        df_noleggio[chiave] = df_noleggio[chiave].astype(str).str.strip().str.upper()
        df_pobs[chiave] = df_pobs[chiave].astype(str).str.strip().str.upper()
        processing_log.append("[OK] Data cleaning completed")

        # Find new records
        processing_log.append("[INFO] Comparing records to find new POBS IDs...")
        nuovi = df_noleggio[~df_noleggio[chiave].isin(df_pobs[chiave])]

        if nuovi.empty:
            processing_log.append("[INFO] No new records found")
            if excluded_count > 0:
                message = f'No new POBS IDs to add. All POBS IDs from Noleggio file already exist in POBS file. Found {excluded_count} records with Resolved-Rejected status that were excluded.'
            else:
                message = 'No new POBS IDs to add. All POBS IDs from Noleggio file already exist in POBS file.'
            result = {
                'success': True,
                'new_records_count': 0,
                'excluded_resolved_rejected_count': excluded_count,
                'message': message,
                'preview_data': [],
                'processing_log': processing_log
            }
            return result

        processing_log.append(f"[OK] Found {len(nuovi)} new records to process")

        # Get preview columns (A-J + M-X)
        processing_log.append("[INFO] Preparing preview data...")
        col_indices = list(range(0, 10)) + list(range(12, 24))
        anteprima_cols = [nuovi.columns[i] for i in col_indices if i < len(nuovi.columns)]
        processing_log.append(f"[INFO] Using {len(anteprima_cols)} columns for preview")

        # Prepare preview data
        preview_data = []
        for _, row in nuovi.head(50).iterrows():  # First 50 records for preview
            record = {}
            for i, col_name in enumerate(anteprima_cols):
                if i < len(col_indices) and col_indices[i] < len(row):
                    value = row.iloc[col_indices[i]]
                    # Convert NaN and None to null for JSON serialization
                    if pd.isna(value) or value is None:
                        record[col_name] = None
                    else:
                        record[col_name] = value
                else:
                    record[col_name] = None
            preview_data.append(record)

        # Log the verification operation
        log_details = {
            "noleggio_file": os.path.basename(noleggio_path),
            "pobs_file": os.path.basename(pobs_path),
            "noleggio_total_records": len(df_noleggio),
            "pobs_total_records": len(df_pobs),
            "new_records_found": len(nuovi),
            "preview_columns": anteprima_cols
        }

        log_filename = log_pobs_operation(
            operation_name="VERIFY_NEW_RECORDS",
            status="SUCCESS",
            details=log_details
        )

        processing_log.append(f"[OK] Preview data prepared for {len(preview_data)} records")
        processing_log.append("[OK] POBS verification completed successfully")

        result = {
            'success': True,
            'new_records_count': len(nuovi),
            'columns': anteprima_cols,
            'preview_data': preview_data,
            'message': f'Found {len(nuovi)} new records to add.',
            'log_file': log_filename,
            'processing_log': processing_log
        }
        return result

    except Exception as e:
        processing_log.append(f"[ERROR] Operation failed: {str(e)}")

        # Log error
        error_details = {
            "noleggio_file": os.path.basename(noleggio_path) if noleggio_path else "Unknown",
            "pobs_file": os.path.basename(pobs_path) if pobs_path else "Unknown",
            "error_message": str(e)
        }

        log_filename = log_pobs_operation(
            operation_name="VERIFY_NEW_RECORDS",
            status="ERROR",
            details=error_details,
            errors=[str(e)]
        )

        return {
            'success': False,
            'error': str(e),
            'log_file': log_filename,
            'processing_log': processing_log
        }

def verify_new_records_realtime(noleggio_path, pobs_path, session_id=None):
    """
    Verify new records between Noleggio and POBS files with real-time logging
    """
    processing_log = []

    def log_message(message):
        processing_log.append(message)
        if session_id:
            if message.startswith('[INFO]'):
                realtime_logger.log_info(session_id, message[6:])
            elif message.startswith('[OK]'):
                realtime_logger.log_ok(session_id, message[4:])
            elif message.startswith('[WARNING]'):
                realtime_logger.log_warning(session_id, message[9:])
            elif message.startswith('[ERROR]'):
                realtime_logger.log_error(session_id, message[7:])
            else:
                realtime_logger.log(session_id, message)

    try:
        log_message("[INFO] Starting POBS verification process...")
        chiave = "POBS ID"

        # Read files
        log_message(f"[INFO] Reading Noleggio file: {os.path.basename(noleggio_path)}")
        df_noleggio = pd.read_excel(noleggio_path, dtype=str)
        log_message(f"[OK] Loaded {len(df_noleggio)} records from Noleggio file")

        # Apply status filtering for realtime POBS verification
        log_message("[INFO] Applying status filter to exclude 'Resolved-Rejected' / 'Risolto-Rifiutato' records")
        df_noleggio, excluded_count = filter_resolved_rejected_status(df_noleggio, log_message)
        if excluded_count > 0:
            log_message(f"[INFO] After filtering: {len(df_noleggio)} records remaining (excluded {excluded_count} resolved-rejected records)")
        else:
            log_message("[INFO] No records excluded by status filter")

        log_message(f"[INFO] Reading POBS file: {os.path.basename(pobs_path)}")
        df_pobs = pd.read_excel(pobs_path, dtype=str)
        log_message(f"[OK] Loaded {len(df_pobs)} records from POBS file")

        # Check if required column exists
        log_message(f"[INFO] Checking for required column: '{chiave}'")
        if chiave not in df_noleggio.columns:
            log_message(f"[ERROR] Column '{chiave}' not found in Noleggio file")
            if session_id:
                realtime_logger.complete_session(session_id)
            return {
                'success': False,
                'error': f"Column '{chiave}' not found in Noleggio file. Available columns: {list(df_noleggio.columns)}",
                'processing_log': processing_log
            }

        if chiave not in df_pobs.columns:
            log_message(f"[ERROR] Column '{chiave}' not found in POBS file")
            if session_id:
                realtime_logger.complete_session(session_id)
            return {
                'success': False,
                'error': f"Column '{chiave}' not found in POBS file. Available columns: {list(df_pobs.columns)}",
                'processing_log': processing_log
            }

        log_message(f"[OK] Required column '{chiave}' found in both files")

        # Clean and normalize data
        log_message("[INFO] Cleaning and normalizing data...")
        df_noleggio[chiave] = df_noleggio[chiave].astype(str).str.strip().str.upper()
        df_pobs[chiave] = df_pobs[chiave].astype(str).str.strip().str.upper()
        log_message("[OK] Data cleaning completed")

        # Find new records
        log_message("[INFO] Comparing records to find new POBS IDs...")
        nuovi = df_noleggio[~df_noleggio[chiave].isin(df_pobs[chiave])]

        if nuovi.empty:
            log_message("[INFO] No new records found")
            if excluded_count > 0:
                message = f'No new POBS IDs to add. All POBS IDs from Noleggio file already exist in POBS file. Found {excluded_count} records with Resolved-Rejected status that were excluded.'
            else:
                message = 'No new POBS IDs to add. All POBS IDs from Noleggio file already exist in POBS file.'
            result = {
                'success': True,
                'new_records_count': 0,
                'excluded_resolved_rejected_count': excluded_count,
                'message': message,
                'preview_data': [],
                'processing_log': processing_log
            }
            if session_id:
                realtime_logger.store_result(session_id, result)
                realtime_logger.complete_session(session_id)
            return result

        log_message(f"[OK] Found {len(nuovi)} new records to process")

        # Get preview columns (A-J + M-X)
        log_message("[INFO] Preparing preview data...")
        col_indices = list(range(0, 10)) + list(range(12, 24))
        anteprima_cols = [nuovi.columns[i] for i in col_indices if i < len(nuovi.columns)]
        log_message(f"[INFO] Using {len(anteprima_cols)} columns for preview")

        # Prepare preview data
        preview_data = []
        for _, row in nuovi.head(50).iterrows():  # First 50 records for preview
            record = {}
            for i, col_name in enumerate(anteprima_cols):
                if i < len(col_indices) and col_indices[i] < len(row):
                    value = row.iloc[col_indices[i]]
                    # Convert NaN and None to null for JSON serialization
                    if pd.isna(value) or value is None:
                        record[col_name] = None
                    else:
                        record[col_name] = value
                else:
                    record[col_name] = None
            preview_data.append(record)

        # Log the verification operation
        log_details = {
            "noleggio_file": os.path.basename(noleggio_path),
            "pobs_file": os.path.basename(pobs_path),
            "noleggio_total_records": len(df_noleggio),
            "pobs_total_records": len(df_pobs),
            "new_records_found": len(nuovi),
            "preview_columns": anteprima_cols
        }

        log_filename = log_pobs_operation(
            operation_name="VERIFY_NEW_RECORDS",
            status="SUCCESS",
            details=log_details
        )

        log_message(f"[OK] Preview data prepared for {len(preview_data)} records")
        log_message("[OK] POBS verification completed successfully")

        result = {
            'success': True,
            'new_records_count': len(nuovi),
            'columns': anteprima_cols,
            'preview_data': preview_data,
            'message': f'Found {len(nuovi)} new records to add.',
            'log_file': log_filename,
            'processing_log': processing_log
        }

        if session_id:
            realtime_logger.store_result(session_id, result)
            realtime_logger.complete_session(session_id)

        return result

    except Exception as e:
        log_message(f"[ERROR] Operation failed: {str(e)}")

        # Log error
        error_details = {
            "noleggio_file": os.path.basename(noleggio_path) if noleggio_path else "Unknown",
            "pobs_file": os.path.basename(pobs_path) if pobs_path else "Unknown",
            "error_message": str(e)
        }

        log_filename = log_pobs_operation(
            operation_name="VERIFY_NEW_RECORDS",
            status="ERROR",
            details=error_details,
            errors=[str(e)]
        )

        result = {
            'success': False,
            'error': str(e),
            'log_file': log_filename,
            'processing_log': processing_log
        }

        if session_id:
            realtime_logger.store_result(session_id, result)
            realtime_logger.complete_session(session_id)

        return result

def add_new_records_realtime(noleggio_path, pobs_path, session_id=None):
    """
    Add new records to POBS file with real-time logging
    """
    processing_log = []

    def log_message(message):
        processing_log.append(message)
        if session_id:
            if message.startswith('[INFO]'):
                realtime_logger.log_info(session_id, message[6:])
            elif message.startswith('[OK]'):
                realtime_logger.log_ok(session_id, message[4:])
            elif message.startswith('[WARNING]'):
                realtime_logger.log_warning(session_id, message[9:])
            elif message.startswith('[ERROR]'):
                realtime_logger.log_error(session_id, message[7:])
            else:
                realtime_logger.log_info(session_id, message)

    try:
        log_message("[INFO] Starting POBS add new records process...")

        # First verify to get new records (without session_id to avoid conflicts)
        log_message("[INFO] Running verification to find new records...")
        verification_result = verify_new_records_realtime(noleggio_path, pobs_path, None)
        if not verification_result['success']:
            log_message(f"[ERROR] Verification failed: {verification_result.get('error', 'Unknown error')}")
            if session_id:
                realtime_logger.complete_session(session_id)
            return {
                'success': False,
                'message': f'Verification failed: {verification_result.get("error", "Unknown error")}',
                'processing_log': processing_log
            }

        if verification_result.get('new_records_count', 0) == 0:
            log_message("[INFO] No new records found to add")
            excluded_count = verification_result.get('excluded_resolved_rejected_count', 0)
            if excluded_count > 0:
                message = f'No new records to add. All POBS IDs from Noleggio file already exist in POBS file. Found {excluded_count} records with Resolved-Rejected status that were excluded.'
            else:
                message = 'No new records to add. All POBS IDs from Noleggio file already exist in POBS file.'
            result = {
                'success': True,
                'no_changes': True,
                'message': message,
                'excluded_resolved_rejected_count': excluded_count,
                'processing_log': processing_log
            }
            if session_id:
                realtime_logger.store_result(session_id, result)
                realtime_logger.complete_session(session_id)
            return result

        log_message(f"[INFO] Found {verification_result['new_records_count']} new records to add")

        # Read files to get actual new records
        log_message("[INFO] Reading Noleggio file...")
        df_noleggio = pd.read_excel(noleggio_path, dtype=str)
        log_message(f"[OK] Loaded {len(df_noleggio)} records from Noleggio file")

        log_message("[INFO] Reading POBS file...")
        df_pobs = pd.read_excel(pobs_path, dtype=str)
        original_count = len(df_pobs)
        log_message(f"[OK] Original POBS file has {original_count} records")

        # Find new records (same logic as verify function)
        chiave = "POBS ID"
        log_message("[INFO] Finding new records to add...")
        df_noleggio[chiave] = df_noleggio[chiave].astype(str).str.strip().str.upper()
        df_pobs[chiave] = df_pobs[chiave].astype(str).str.strip().str.upper()

        nuovi = df_noleggio[~df_noleggio[chiave].isin(df_pobs[chiave])]
        log_message(f"[OK] Found {len(nuovi)} new records to add")

        # Append new records to POBS
        df_combined = pd.concat([df_pobs, nuovi], ignore_index=True)

        # Create POBS output directory
        pobs_dir = os.path.join('outputs', 'POBS')
        os.makedirs(pobs_dir, exist_ok=True)

        # Save updated file
        output_filename = f"POBS_updated_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = os.path.join(pobs_dir, output_filename)

        log_message("[INFO] Saving updated POBS file...")
        df_combined.to_excel(output_path, index=False)

        final_count = len(df_combined)
        records_added = len(nuovi)

        log_message(f"[OK] Successfully added {records_added} records")
        log_message(f"[OK] Final POBS file has {final_count} records")
        log_message(f"[OK] Updated file saved: {output_filename}")

        result = {
            'success': True,
            'message': f'Successfully added {records_added} new records',
            'records_added': records_added,
            'original_count': original_count,
            'final_count': final_count,
            'download_file': output_filename,
            'processing_log': processing_log
        }

        if session_id:
            realtime_logger.store_result(session_id, result)
            realtime_logger.complete_session(session_id)

        return result

    except Exception as e:
        log_message(f"[ERROR] Operation failed: {str(e)}")
        result = {
            'success': False,
            'error': str(e),
            'processing_log': processing_log
        }
        if session_id:
            realtime_logger.store_result(session_id, result)
            realtime_logger.complete_session(session_id)
        return result

def add_new_records(noleggio_path, pobs_path, output_dir):
    """
    Add new records to POBS file
    Converted from aggiungi_nuovi function
    """
    processing_log = []

    try:
        processing_log.append("[INFO] Starting POBS add new records process...")

        # First verify to get new records
        processing_log.append("[INFO] Running verification to find new records...")
        verification_result = verify_new_records(noleggio_path, pobs_path)
        if not verification_result['success']:
            processing_log.append(f"[ERROR] Verification failed: {verification_result.get('error', 'Unknown error')}")
            return {
                'success': False,
                'message': f'Verification failed: {verification_result.get("error", "Unknown error")}',
                'processing_log': processing_log
            }

        if verification_result['new_records_count'] == 0:
            processing_log.append("[INFO] No new records found to add")
            excluded_count = verification_result.get('excluded_resolved_rejected_count', 0)
            if excluded_count > 0:
                message = f'No new records found to add. All POBS IDs from Noleggio file already exist in POBS file. Found {excluded_count} records with Resolved-Rejected status that were excluded.'
            else:
                message = 'No new records found to add. All POBS IDs from Noleggio file already exist in POBS file.'
            return {
                'success': True,
                'no_changes': True,
                'message': message,
                'records_added': 0,
                'excluded_resolved_rejected_count': excluded_count,
                'total_noleggio_records': len(pd.read_excel(noleggio_path, dtype=str)),
                'total_pobs_records': len(pd.read_excel(pobs_path, dtype=str)),
                'processing_log': processing_log
            }

        processing_log.append(f"[OK] Found {verification_result['new_records_count']} new records to add")

        # Re-read files for processing
        processing_log.append("[INFO] Reading files for processing...")
        chiave = "POBS ID"
        df_noleggio = pd.read_excel(noleggio_path, dtype=str)
        df_pobs = pd.read_excel(pobs_path, dtype=str)

        processing_log.append("[INFO] Cleaning and normalizing data...")
        df_noleggio[chiave] = df_noleggio[chiave].astype(str).str.strip().str.upper()
        df_pobs[chiave] = df_pobs[chiave].astype(str).str.strip().str.upper()

        nuovi = df_noleggio[~df_noleggio[chiave].isin(df_pobs[chiave])]
        processing_log.append(f"[OK] Identified {len(nuovi)} records to add")

        # Create backup
        processing_log.append("[INFO] Creating backup of original POBS file...")
        cartella_backup = os.path.join(output_dir, "Backup")
        os.makedirs(cartella_backup, exist_ok=True)
        backup_filename = f"{os.path.splitext(os.path.basename(pobs_path))[0]}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        backup_file = os.path.join(cartella_backup, backup_filename)
        shutil.copy2(pobs_path, backup_file)
        processing_log.append(f"[OK] Backup created: {backup_filename}")

        # Load workbook for modification
        processing_log.append("[INFO] Loading POBS workbook for modification...")
        wb = load_workbook(pobs_path)
        ws = wb.active
        headers_pobs = [cell.value for cell in ws[1]]
        tot_colonne = len(headers_pobs)
        processing_log.append(f"[OK] Workbook loaded with {tot_colonne} columns")

        # Column mapping (same as original)
        processing_log.append("[INFO] Setting up column mapping...")
        mappa = {
            0: 0, 1: 1, 2: 2, 3: 3, 4: 4,
            5: 5, 6: 6, 7: 7, 8: 8, 9: 9,   # A-J
            12: 10, 13: 11, 14: 12, 15: 13,
            16: 14, 17: 15, 18: 16, 19: 17,
            20: 18, 21: 19, 23: 20           # M→U
        }

        processing_log.append("[INFO] Adding new records to POBS file...")
        nuovi_id = []
        records_processed = 0
        for _, row in nuovi.iterrows():
            records_processed += 1
            if records_processed % 10 == 0:
                processing_log.append(f"[INFO] Processed {records_processed}/{len(nuovi)} records...")
            nuova_riga = [None] * tot_colonne
            for col_noleggio, col_pobs in mappa.items():
                if col_noleggio < len(row):
                    nuova_riga[col_pobs] = row.iloc[col_noleggio]

            # Column Y (index 24) = "IN GESTIONE"
            if 24 < tot_colonne:
                nuova_riga[24] = "IN GESTIONE"

            ws.append(nuova_riga)
            nuovi_id.append(row[chiave])

        processing_log.append(f"[OK] Successfully added {len(nuovi_id)} records to worksheet")

        # Create POBS output directory
        processing_log.append("[INFO] Creating output directory...")
        pobs_dir = os.path.join(output_dir, "POBS")
        os.makedirs(pobs_dir, exist_ok=True)

        # Save updated file
        processing_log.append("[INFO] Saving updated POBS file...")
        updated_filename = f"{os.path.splitext(os.path.basename(pobs_path))[0]}_updated_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        updated_file = os.path.join(pobs_dir, updated_filename)
        wb.save(updated_file)
        processing_log.append(f"[OK] Updated file saved: {updated_filename}")

        # Create log
        log_file = os.path.join(cartella_backup, "aggiornamenti.log")
        with open(log_file, "a", encoding="utf-8") as log:
            log.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] [AGGIUNTA RIGHE] "
                      f"POBS: {pobs_path} "
                      f"| Backup: {backup_file} "
                      f"| Noleggio: {noleggio_path} "
                      f"| Righe aggiunte: {len(nuovi_id)} "
                      f"| Nuovi POBS ID: {', '.join(nuovi_id)}\n")

        # Log the operation using the new logging system
        log_details = {
            "noleggio_file": os.path.basename(noleggio_path),
            "pobs_file": os.path.basename(pobs_path),
            "records_added": len(nuovi_id),
            "backup_file": backup_filename,
            "updated_file": updated_filename,
            "new_pobs_ids": nuovi_id[:10],  # First 10 for logging
            "total_new_records": len(nuovi_id)
        }

        log_filename = log_pobs_operation(
            operation_name="ADD_NEW_RECORDS",
            status="SUCCESS",
            details=log_details,
            files_created=[updated_filename, backup_filename]
        )

        processing_log.append("[OK] POBS add new records operation completed successfully")

        return {
            'success': True,
            'message': f'Successfully added {len(nuovi_id)} new records.',
            'records_added': len(nuovi_id),
            'backup_file': backup_filename,
            'updated_file': updated_filename,
            'new_pobs_ids': nuovi_id[:10],  # First 10 IDs
            'download_file': updated_filename,
            'log_file': log_filename,
            'processing_log': processing_log
        }

    except Exception as e:
        processing_log.append(f"[ERROR] Operation failed: {str(e)}")

        # Log error
        error_details = {
            "noleggio_file": os.path.basename(noleggio_path) if noleggio_path else "Unknown",
            "pobs_file": os.path.basename(pobs_path) if pobs_path else "Unknown",
            "error_message": str(e)
        }

        log_filename = log_pobs_operation(
            operation_name="ADD_NEW_RECORDS",
            status="ERROR",
            details=error_details,
            errors=[str(e)]
        )

        return {
            'success': False,
            'error': str(e),
            'log_file': log_filename,
            'processing_log': processing_log
        }

def update_imei_data_realtime(pobs_path, master_path, template_path, session_id=None, custom_name=None):
    """
    Update IMEI data from masterfile with real-time logging
    """
    processing_log = []

    def log_message(message):
        processing_log.append(message)
        if session_id:
            if message.startswith('[INFO]'):
                realtime_logger.log_info(session_id, message[6:])
            elif message.startswith('[OK]'):
                realtime_logger.log_ok(session_id, message[4:])
            elif message.startswith('[WARNING]'):
                realtime_logger.log_warning(session_id, message[9:])
            elif message.startswith('[ERROR]'):
                realtime_logger.log_error(session_id, message[7:])
            else:
                realtime_logger.log_info(session_id, message)

    try:
        log_message("[INFO] Starting IMEI data update process...")

        # Read files
        log_message("[INFO] Reading POBS file...")
        df_pobs = pd.read_excel(pobs_path)
        log_message(f"[OK] POBS file loaded with {len(df_pobs)} records")

        log_message("[INFO] Reading master file...")
        df_master = pd.read_excel(master_path)
        log_message(f"[OK] Master file loaded with {len(df_master)} records")

        log_message("[INFO] Reading template file...")
        df_template = pd.read_excel(template_path)
        log_message(f"[OK] Template file loaded with {len(df_template)} records")

        # Check if IMEI columns exist
        pobs_imei_col = None
        master_imei_col = None

        for col in df_pobs.columns:
            if 'imei' in col.lower():
                pobs_imei_col = col
                break

        for col in df_master.columns:
            if 'imei' in col.lower():
                master_imei_col = col
                break

        if not pobs_imei_col:
            log_message("[ERROR] IMEI column not found in POBS file")
            if session_id:
                realtime_logger.complete_session(session_id)
            return {
                'success': False,
                'error': 'IMEI column not found in POBS file',
                'processing_log': processing_log
            }

        if not master_imei_col:
            log_message("[ERROR] IMEI column not found in master file")
            if session_id:
                realtime_logger.complete_session(session_id)
            return {
                'success': False,
                'error': 'IMEI column not found in master file',
                'processing_log': processing_log
            }

        log_message(f"[INFO] Found IMEI columns: POBS='{pobs_imei_col}', Master='{master_imei_col}'")

        # Update IMEI data
        log_message("[INFO] Updating IMEI data...")
        df_result = df_template.copy()
        template_warnings = []

        # Create a mapping from master file
        master_mapping = dict(zip(df_master[master_imei_col].astype(str), df_master.index))

        updated_count = 0
        empty_cell_count = 0

        for idx, row in df_pobs.iterrows():
            imei_value = str(row[pobs_imei_col])
            if imei_value in master_mapping:
                master_idx = master_mapping[imei_value]
                # Copy data from master to result and check for empty cells
                for col in df_master.columns:
                    if col in df_result.columns:
                        value = df_master.loc[master_idx, col]
                        df_result.loc[updated_count, col] = value
                        # Check for empty cells
                        if value is None or value == '' or (isinstance(value, str) and value.strip() == ''):
                            empty_cell_count += 1
                updated_count += 1

        # Add warning if empty cells found
        if empty_cell_count > 0:
            warning_msg = f"⚠️ Warning: Found {empty_cell_count} empty cells in template columns. Please review the output file for missing data."
            log_message(f"[WARNING] {warning_msg}")
            template_warnings.append(warning_msg)

        log_message(f"[OK] Updated {updated_count} IMEI records")

        # Generate output filename
        # Create IMEI HUB output directory
        imei_hub_dir = os.path.join('outputs', 'IMEI HUB')
        os.makedirs(imei_hub_dir, exist_ok=True)

        if custom_name:
            output_filename = custom_name if custom_name.endswith('.xlsx') else f"{custom_name}.xlsx"
        else:
            output_filename = f"IMEI_Updated_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        # Save updated file
        output_path = os.path.join(imei_hub_dir, output_filename)
        log_message("[INFO] Saving updated file...")
        df_result.to_excel(output_path, index=False)
        log_message(f"[OK] Updated file saved: {output_filename}")

        result_message = f'Successfully updated {updated_count} IMEI records'
        if template_warnings:
            result_message += f'. {template_warnings[0]}'

        result = {
            'success': True,
            'message': result_message,
            'records_updated': updated_count,
            'download_file': output_filename,
            'download_files': [output_filename],
            'processing_log': processing_log,
            'warnings': template_warnings
        }

        if session_id:
            realtime_logger.store_result(session_id, result)
            realtime_logger.complete_session(session_id)

        return result

    except Exception as e:
        log_message(f"[ERROR] Operation failed: {str(e)}")
        result = {
            'success': False,
            'error': str(e),
            'processing_log': processing_log
        }
        if session_id:
            realtime_logger.store_result(session_id, result)
            realtime_logger.complete_session(session_id)
        return result

def update_imei_data(pobs_path, master_path, template_path, output_dir, custom_name=None):
    """
    Update IMEI data from masterfile with enhanced formatting and custom naming
    Updated from script 2 with improved output handling
    """
    processing_log = []

    try:
        processing_log.append("[INFO] Starting POBS IMEI data update process...")

        # Create backup
        processing_log.append("[INFO] Creating backup of POBS file...")
        cartella_backup = os.path.join(output_dir, "Backup")
        os.makedirs(cartella_backup, exist_ok=True)
        backup_filename = f"{os.path.splitext(os.path.basename(pobs_path))[0]}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        backup_file = os.path.join(cartella_backup, backup_filename)
        shutil.copy2(pobs_path, backup_file)
        processing_log.append(f"[OK] Backup created: {backup_filename}")

        # Load masterfile "PER STOPRIPARO" sheet
        processing_log.append(f"[INFO] Loading master file: {os.path.basename(master_path)}")
        df_master = pd.read_excel(master_path, dtype=str, sheet_name="PER STOPRIPARO")
        df_master = df_master.iloc[:, [1, 2, 7]]  # B=GUID, C=IMEI, H=Data spedizione
        df_master.columns = ["GUID", "IMEI", "DATA_SPED"]
        df_master["GUID"] = df_master["GUID"].astype(str).str.strip().str.upper()
        processing_log.append(f"[OK] Loaded {len(df_master)} records from master file")

        processing_log.append("[INFO] Creating GUID to data mapping...")
        guid_to_data = dict(zip(df_master["GUID"], zip(df_master["IMEI"], df_master["DATA_SPED"])))
        processing_log.append(f"[OK] Created mapping for {len(guid_to_data)} GUIDs")

        # Load POBS workbook
        processing_log.append(f"[INFO] Loading POBS workbook: {os.path.basename(pobs_path)}")
        wb = load_workbook(pobs_path)
        ws = wb.active
        processing_log.append("[OK] POBS workbook loaded successfully")

        # Column indices (same as original)
        col_guid = 8   # H = GUID
        col_imei = 10  # J = IMEI
        col_data = 23  # W = Data spedizione
        col_stato = 25 # Y = Stato

        aggiornati = 0
        aggiornati_id = []
        righe_template = []
        data_sped_finale = None

        processing_log.append("[INFO] Starting IMEI data updates...")
        total_rows = ws.max_row - 1  # Excluding header
        rows_processed = 0

        # Update records with improved formatting
        for row in ws.iter_rows(min_row=2):
            rows_processed += 1
            if rows_processed % 100 == 0:
                processing_log.append(f"[INFO] Processed {rows_processed}/{total_rows} rows...")
            guid_val = str(row[col_guid-1].value).strip().upper() if row[col_guid-1].value else None
            if guid_val and guid_val in guid_to_data:
                imei_val, data_sped = guid_to_data[guid_val]

                # Update IMEI as number with proper formatting
                if imei_val and imei_val.strip():
                    try:
                        imei_num = int(imei_val)
                        cell_imei = ws.cell(row=row[0].row, column=col_imei, value=imei_num)
                        cell_imei.number_format = numbers.FORMAT_NUMBER
                    except:
                        ws.cell(row=row[0].row, column=col_imei, value=imei_val)

                # Update shipping date with proper formatting
                if data_sped and str(data_sped).strip():
                    try:
                        data_conv = pd.to_datetime(data_sped, errors="coerce", dayfirst=True)
                        if pd.notnull(data_conv):
                            cell_data = ws.cell(row=row[0].row, column=col_data, value=data_conv)
                            cell_data.number_format = "DD/MM/YYYY"
                        else:
                            ws.cell(row=row[0].row, column=col_data, value=data_sped)
                    except:
                        ws.cell(row=row[0].row, column=col_data, value=data_sped)

                # Update status
                ws.cell(row=row[0].row, column=col_stato, value="SPEDITO")

                aggiornati += 1
                aggiornati_id.append(guid_val)

                # Collect data for template (columns A-J)
                valori = [row[i].value for i in range(10)]
                righe_template.append(valori)
                if not data_sped_finale and data_sped:
                    data_sped_finale = data_sped

        processing_log.append(f"[OK] Updated {aggiornati} records with IMEI data")

        # Save updated POBS file and create downloadable copy
        processing_log.append("[INFO] Saving updated POBS file...")
        wb.save(pobs_path)
        processing_log.append("[OK] Original POBS file updated and saved")

        # Create downloadable copy of updated POBS file
        pobs_output_dir = os.path.join(output_dir, "POBS")
        os.makedirs(pobs_output_dir, exist_ok=True)

        pobs_updated_filename = f"{os.path.splitext(os.path.basename(pobs_path))[0]}_updated_with_IMEI_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        pobs_updated_path = os.path.join(pobs_output_dir, pobs_updated_filename)
        wb.save(pobs_updated_path)
        processing_log.append(f"[OK] Updated POBS file saved for download: {pobs_updated_filename}")

        # Generate IMEI HUB file if there are updated records
        imei_hub_path = None
        template_warnings = []
        if righe_template:
            processing_log.append(f"[INFO] Generating IMEI HUB file for {len(righe_template)} updated records...")
            wb_template = load_workbook(template_path)
            ws_template = wb_template.active

            # Check for empty cells in template columns and add warning
            empty_cell_count = 0
            for valori in righe_template:
                ws_template.append(valori)
                # Format IMEI column (column 10) as number
                last_row = ws_template.max_row
                imei_cell = ws_template.cell(row=last_row, column=10)
                try:
                    imei_cell.value = int(imei_cell.value)
                    imei_cell.number_format = numbers.FORMAT_NUMBER
                except:
                    pass

                # Check for empty cells in the row (columns A-J, indices 0-9)
                for col_idx, val in enumerate(valori):
                    if val is None or val == '' or (isinstance(val, str) and val.strip() == ''):
                        empty_cell_count += 1

            if empty_cell_count > 0:
                warning_msg = f"⚠️ Warning: Found {empty_cell_count} empty cells in template columns. Please review the output file for missing data."
                processing_log.append(f"[WARNING] {warning_msg}")
                template_warnings.append(warning_msg)

            # Create IMEI HUB directory
            imei_hub_dir = os.path.join(output_dir, "IMEI HUB")
            os.makedirs(imei_hub_dir, exist_ok=True)

            # Generate custom filename if provided, otherwise prompt-style default
            if custom_name:
                imei_hub_filename = custom_name
            else:
                imei_hub_filename = f"IMEI_HUB_{datetime.now().strftime('%Y%m%d')}.xlsx"

            imei_hub_path = os.path.join(imei_hub_dir, imei_hub_filename)
            wb_template.save(imei_hub_path)
            processing_log.append(f"[OK] IMEI HUB file saved: {imei_hub_filename}")
        else:
            processing_log.append("[INFO] No records updated - IMEI HUB file not generated")

        # Create log
        log_file = os.path.join(cartella_backup, "aggiornamenti.log")
        with open(log_file, "a", encoding="utf-8") as log:
            log.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] [AGGIORNAMENTO IMEI+STATO+DATA+EXPORT] "
                      f"POBS: {pobs_path} "
                      f"| Backup: {backup_file} "
                      f"| Masterfile: {master_path} "
                      f"| Righe aggiornate: {aggiornati} "
                      f"| File IMEI HUB: {imei_hub_filename if imei_hub_filename else 'Nessun file generato'} "
                      f"| GUID aggiornati: {', '.join(aggiornati_id[:10])}{'...' if len(aggiornati_id)>10 else ''}\n")

        # Log the operation using the new logging system
        log_details = {
            "pobs_file": os.path.basename(pobs_path),
            "master_file": os.path.basename(master_path),
            "template_file": os.path.basename(template_path) if template_path else "None",
            "records_updated": aggiornati,
            "backup_file": backup_filename,
            "imei_hub_file": os.path.basename(imei_hub_path) if imei_hub_path else None,
            "updated_guids": aggiornati_id[:10],  # First 10 for logging
            "custom_name": custom_name or "Default"
        }

        files_created = [backup_filename]
        if imei_hub_path:
            files_created.append(os.path.basename(imei_hub_path))

        log_filename = log_pobs_operation(
            operation_name="UPDATE_IMEI_DATA",
            status="SUCCESS",
            details=log_details,
            files_created=files_created
        )

        processing_log.append("[OK] POBS IMEI data update operation completed successfully")

        result_message = f'Successfully updated {aggiornati} records. Generated updated POBS file and IMEI HUB file.'
        if template_warnings:
            result_message += f' {template_warnings[0]}'

        # Prepare download files list
        download_files = []
        if aggiornati > 0:  # Only include updated POBS if records were actually updated
            download_files.append(pobs_updated_filename)
        if imei_hub_path:
            download_files.append(os.path.basename(imei_hub_path))

        return {
            'success': True,
            'message': result_message,
            'records_updated': aggiornati,
            'backup_file': backup_filename,
            'updated_pobs_file': pobs_updated_filename if aggiornati > 0 else None,
            'imei_hub_file': os.path.basename(imei_hub_path) if imei_hub_path else None,
            'updated_guids': aggiornati_id[:10],  # First 10 GUIDs
            'download_files': download_files,
            'log_file': log_filename,
            'processing_log': processing_log,
            'warnings': template_warnings
        }

    except Exception as e:
        processing_log.append(f"[ERROR] Operation failed: {str(e)}")

        # Log error
        error_details = {
            "pobs_file": os.path.basename(pobs_path) if pobs_path else "Unknown",
            "master_file": os.path.basename(master_path) if master_path else "Unknown",
            "template_file": os.path.basename(template_path) if template_path else "Unknown",
            "custom_name": custom_name or "Default",
            "error_message": str(e)
        }

        log_filename = log_pobs_operation(
            operation_name="UPDATE_IMEI_DATA",
            status="ERROR",
            details=error_details,
            errors=[str(e)]
        )

        return {
            'success': False,
            'error': str(e),
            'log_file': log_filename,
            'processing_log': processing_log
        }