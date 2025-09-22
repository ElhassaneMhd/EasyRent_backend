"""
Tracking Service - Converted from tracking_manager.py
Removes Tkinter GUI and converts to web service functions
"""

import openpyxl
import xlwt
import xlrd
import os
import re
import uuid
import pandas as pd
from datetime import datetime
from collections import defaultdict
from .logger_service import log_tracking_operation
from .realtime_logger import realtime_logger

def generate_upload_gsped(pobs_path, masterfile_path, output_dir):
    """
    Generate Upload Gsped file
    Converted from genera_upload function
    """
    processing_log = []

    try:
        processing_log.append("[INFO] Starting Upload Gsped generation process...")

        # Create GSPED output directory
        processing_log.append("[INFO] Creating GSPED output directory...")
        gsped_dir = os.path.join(output_dir, "GSPED")
        os.makedirs(gsped_dir, exist_ok=True)
        processing_log.append("[OK] GSPED directory ready")

        # Generate progressive filename
        processing_log.append("[INFO] Generating output filename...")
        oggi = datetime.now().strftime("%Y%m%d")
        pattern = re.compile(rf"^Upload Gsped_{oggi}_(\d+)\.xls$")

        progressivi = []
        for fname in os.listdir(gsped_dir):
            match = pattern.match(fname)
            if match:
                progressivi.append(int(match.group(1)))
        progressivo = max(progressivi) + 1 if progressivi else 1
        output_filename = f"Upload Gsped_{oggi}_{progressivo:02d}.xls"
        output_path = os.path.join(gsped_dir, output_filename)
        processing_log.append(f"[OK] Output filename: {output_filename}")

        # Load masterfile and extract GUIDs from "PER STOPRIPARO" sheet
        processing_log.append(f"[INFO] Loading master file: {os.path.basename(masterfile_path)}")
        master_wb = openpyxl.load_workbook(masterfile_path, data_only=True)
        if "PER STOPRIPARO" not in master_wb.sheetnames:
            processing_log.append("[ERROR] MasterFile does not contain 'PER STOPRIPARO' sheet")
            raise Exception("MasterFile does not contain 'PER STOPRIPARO' sheet.")
        master_sheet = master_wb["PER STOPRIPARO"]
        processing_log.append("[OK] Master file loaded successfully")

        processing_log.append("[INFO] Extracting GUIDs from master file...")
        master_guids = set()
        for row_idx, row in enumerate(master_sheet.iter_rows(values_only=True), start=1):
            if row_idx == 1:  # Skip header
                continue
            guid = row[1]  # Column B
            if guid:
                master_guids.add(str(guid).strip())
        processing_log.append(f"[OK] Extracted {len(master_guids)} GUIDs from master file")

        # Load POBS file
        processing_log.append(f"[INFO] Loading POBS file: {os.path.basename(pobs_path)}")
        pobs_wb = openpyxl.load_workbook(pobs_path, data_only=True)
        pobs_sheet = pobs_wb.active
        processing_log.append(f"[OK] POBS file loaded with {pobs_sheet.max_row} rows")

        # Load template for headers
        processing_log.append("[INFO] Loading template headers...")
        try:
            template_wb = xlrd.open_workbook("Gsped_template_excel.xls")
            template_sheet = template_wb.sheet_by_index(0)
            headers = template_sheet.row_values(0)
            processing_log.append("[OK] Template headers loaded from file")
        except FileNotFoundError:
            # Default headers if template not found
            headers = [
                "CD_SOCIETA", "CD_DEPOSITO", "CD_UTENTE", "INDIRIZZO", "CIVICO", "LOCALITA",
                "CAP", "PROVINCIA", "CD_PRODOTTO", "DS_PRODOTTO",
                "QT_ORDINATA", "QT_EVASA", "QT_DA_EVADERE", "QT_SPEDITA", "FG_SPEDIRE",
                "NOTE1", "NOTE2", "RIFERIMENTO_ESTERNO",
                "CAMPO_LIBERO1", "CAMPO_LIBERO2", "CAMPO_LIBERO3", "CAMPO_LIBERO4",
                "CAMPO_LIBERO5", "CAMPO_LIBERO6", "RAGIONE_SOCIALE", "EMAIL", "TELEFONO"
            ]
            processing_log.append("[INFO] Using default template headers")

        # Create new XLS workbook
        processing_log.append("[INFO] Creating output workbook...")
        new_wb = xlwt.Workbook()
        new_ws = new_wb.add_sheet("Upload Gsped")

        # Write headers
        for col, header in enumerate(headers):
            new_ws.write(0, col, header)
        processing_log.append(f"[OK] Headers written ({len(headers)} columns)")

        # Process POBS data
        processing_log.append("[INFO] Processing POBS data for Gsped mapping...")
        mapped_rows = []
        r_values = []
        processed_count = 0

        for row_idx, row in enumerate(pobs_sheet.iter_rows(values_only=True), start=1):
            if row_idx == 1:  # Skip header
                continue

            guid = row[7] if len(row) > 7 else None  # Column H (index 7)
            if guid and str(guid).strip() in master_guids:
                if processed_count % 50 == 0 and processed_count > 0:
                    processing_log.append(f"[INFO] Processed {processed_count} matching records...")
                # Map POBS data to Gsped format (same mapping as original)
                mapped_row = [
                    "392369",           # CD_SOCIETA
                    "",                 # CD_DEPOSITO
                    row[11] if len(row) > 11 else "",  # CD_UTENTE (col L)
                    row[13] if len(row) > 13 else "",  # INDIRIZZO (col N)
                    row[15] if len(row) > 15 else "",  # CIVICO (col P)
                    row[14] if len(row) > 14 else "",  # LOCALITA (col O)
                    row[16] if len(row) > 16 else "",  # CAP (col Q)
                    row[17] if len(row) > 17 else "",  # PROVINCIA (col R)
                    row[20] if len(row) > 20 else "",  # CD_PRODOTTO (col U)
                    row[19] if len(row) > 19 else "",  # DS_PRODOTTO (col T)
                    "2",                # QT_ORDINATA
                    "0",                # QT_EVASA
                    "0",                # QT_DA_EVADERE
                    "1",                # QT_SPEDITA
                    "1",                # FG_SPEDIRE
                    "",                 # NOTE1
                    "",                 # NOTE2
                    row[6] if len(row) > 6 else "",    # RIFERIMENTO_ESTERNO (col G)
                    "",                 # CAMPO_LIBERO1
                    "",                 # CAMPO_LIBERO2
                    "",                 # CAMPO_LIBERO3
                    "",                 # CAMPO_LIBERO4
                    "",                 # CAMPO_LIBERO5
                    "",                 # CAMPO_LIBERO6
                    "VODAFONE EASYRENT", # RAGIONE_SOCIALE
                    "",                 # EMAIL
                    ""                  # TELEFONO
                ]

                # Convert None values to empty strings and strip
                cleaned_row = tuple("" if v is None else str(v).strip() for v in mapped_row)
                mapped_rows.append(cleaned_row)
                r_values.append("" if row[6] is None else str(row[6]).strip())
                processed_count += 1

        processing_log.append(f"[OK] Mapped {len(mapped_rows)} records from {processed_count} matching GUIDs")

        # Remove duplicates
        processing_log.append("[INFO] Removing duplicate records...")
        unique_rows = list(dict.fromkeys(mapped_rows))
        n_duplicates = len(mapped_rows) - len(unique_rows)
        processing_log.append(f"[OK] Removed {n_duplicates} duplicate records, {len(unique_rows)} unique records remaining")

        # Count duplicates by R value (RIFERIMENTO_ESTERNO)
        dup_counter = defaultdict(int)
        seen = set()
        for row, r_val in zip(mapped_rows, r_values):
            if row in seen:
                dup_counter[r_val] += 1
            else:
                seen.add(row)

        # Write unique rows to file
        processing_log.append("[INFO] Writing data to output file...")
        row_out = 1
        for mapped_row in unique_rows:
            for col, value in enumerate(mapped_row):
                new_ws.write(row_out, col, value)
            row_out += 1

        new_wb.save(output_path)
        processing_log.append(f"[OK] Output file saved: {output_filename}")
        processing_log.append("[OK] Upload Gsped generation completed successfully")

        # Create structured log using new logging system
        log_details = {
            "pobs_file": os.path.basename(pobs_path),
            "masterfile": os.path.basename(masterfile_path),
            "generated_file": output_filename,
            "total_rows_processed": processed_count,
            "total_rows_mapped": len(mapped_rows),
            "unique_rows": len(unique_rows),
            "duplicate_rows_removed": n_duplicates,
            "duplicate_details": dict(dup_counter) if n_duplicates > 0 else {},
            "master_guids_found": len(master_guids),
            "output_directory": gsped_dir
        }

        log_filename = log_tracking_operation(
            operation_name="GENERATE_UPLOAD_GSPED",
            status="SUCCESS",
            details=log_details,
            files_created=[output_filename]
        )

        # Keep legacy log for backward compatibility
        log_path = os.path.join(gsped_dir, "Upload_Gsped_log.txt")
        with open(log_path, "a", encoding="utf-8") as f:
            f.write("="*60 + "\n")
            f.write(f"Execution: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Generated file: {output_filename}\n")
            f.write(f"Total rows processed: {processed_count}\n")
            f.write(f"Total rows mapped: {len(mapped_rows)}\n")
            f.write(f"Unique rows: {len(unique_rows)}\n")
            f.write(f"Duplicate rows removed: {n_duplicates}\n\n")
            if n_duplicates > 0:
                f.write("Duplicate details by column R:\n")
                for key, val in dup_counter.items():
                    f.write(f"- {key}: {val}\n")
            f.write("\n")

        return {
            'success': True,
            'message': f'Successfully generated Gsped file with {len(unique_rows)} unique records',
            'output_file': output_filename,
            'processed_count': processed_count,
            'total_rows': len(mapped_rows),
            'unique_rows': len(unique_rows),
            'duplicates_removed': n_duplicates,
            'duplicate_details': dict(dup_counter),
            'download_file': output_filename,
            'log_file': log_filename,
            'processing_log': processing_log
        }

    except Exception as e:
        processing_log.append(f"[ERROR] Operation failed: {str(e)}")

        # Log error
        error_details = {
            "pobs_file": os.path.basename(pobs_path) if pobs_path else "Unknown",
            "masterfile": os.path.basename(masterfile_path) if masterfile_path else "Unknown",
            "error_message": str(e),
            "output_directory": output_dir
        }

        log_filename = log_tracking_operation(
            operation_name="GENERATE_UPLOAD_GSPED",
            status="ERROR",
            details=error_details,
            errors=[str(e)]
        )

        return {
            'success': False,
            'error': str(e),
            'log_file': log_filename,
            'processing_log': processing_log
        }

def update_tracking_data(pobs_path, trasporti_path, masterfile_path, output_dir, custom_name=None):
    """
    Update tracking data in POBS and generate TRACKING RADAR with masterfile integration
    Enhanced version with backup and custom naming from script 3
    """
    processing_log = []

    try:
        processing_log.append("[INFO] Starting tracking data update process...")
        # Create backup first
        backup_dir = os.path.join(output_dir, "BACKUP_POBS")
        os.makedirs(backup_dir, exist_ok=True)
        backup_filename = f"{os.path.splitext(os.path.basename(pobs_path))[0]}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        backup_path = os.path.join(backup_dir, backup_filename)
        import shutil
        shutil.copy2(pobs_path, backup_path)

        # Load POBS file
        pobs_wb = openpyxl.load_workbook(pobs_path)
        pobs_sheet = pobs_wb.active

        # Load transport file (CSV or Excel)
        if trasporti_path.lower().endswith(".csv"):
            trasporti_df = pd.read_csv(trasporti_path, dtype=str, sep=None, engine="python")
        else:
            trasporti_df = pd.read_excel(trasporti_path, dtype=str)

        trasporti_df = trasporti_df.fillna("")

        # Check required columns
        if "Riferimento alfanumerico" not in trasporti_df.columns or "N. sped." not in trasporti_df.columns:
            raise Exception("Transport file missing required columns: 'Riferimento alfanumerico' or 'N. sped.'")

        # Create tracking mapping dictionary - clean values from transport file
        def clean_tracking_value(val):
            """Remove formula-like formatting from tracking values"""
            if pd.isna(val):
                return ""
            val_str = str(val).strip()
            # Remove Excel formula formatting like ="value"
            if val_str.startswith('="') and val_str.endswith('"'):
                val_str = val_str[2:-1]
            # Remove any remaining quotes
            val_str = val_str.replace('"', '').strip()
            return val_str

        mapping_tracking = {}
        for _, row in trasporti_df.iterrows():
            ref = str(row["Riferimento alfanumerico"]).strip() if pd.notna(row["Riferimento alfanumerico"]) else ""
            tracking = clean_tracking_value(row["N. sped."])
            if ref:
                mapping_tracking[ref] = tracking

        # Load MasterFile for shipping dates
        master_dates = {}
        if masterfile_path:
            master_wb = openpyxl.load_workbook(masterfile_path, data_only=True)
            if "PER STOPRIPARO" in master_wb.sheetnames:
                sheet = master_wb["PER STOPRIPARO"]
                for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    if i == 1:  # Skip header
                        continue
                    guid = str(row[1]).strip() if row[1] else None
                    data_sped = row[2]  # Column "DATA SPEDIZIONE/BOLLA"
                    if guid:
                        master_dates[guid] = data_sped

        # Find required columns in POBS
        headers = [c.value for c in pobs_sheet[1]]
        if "TRACKING - LDV TNT" not in headers or "DATA SPEDIZIONE" not in headers:
            raise Exception("Required columns not found in POBS.")
        tracking_col_idx = headers.index("TRACKING - LDV TNT") + 1
        data_sped_col_idx = headers.index("DATA SPEDIZIONE") + 1

        updates = 0
        updated_rows = []

        # Update POBS with tracking numbers and shipping dates
        for i, row in enumerate(pobs_sheet.iter_rows(min_row=2, values_only=False), start=2):
            guid_pobs = row[6].value  # Column G
            if guid_pobs:
                guid_str = str(guid_pobs).strip()
                changed = False

                # Update tracking - clean value to avoid formula-like formatting
                if guid_str in mapping_tracking:
                    tracking_value = mapping_tracking[guid_str].strip()
                    # Remove any quotes or special characters that might cause Excel formula issues
                    tracking_value = tracking_value.replace('="', '').replace('"', '').strip()
                    pobs_sheet.cell(row=i, column=tracking_col_idx, value=tracking_value)
                    changed = True

                # Update shipping date from masterfile
                if guid_str in master_dates:
                    date_val = master_dates[guid_str]
                    if isinstance(date_val, datetime):
                        from datetime import date
                        only_date = date(date_val.year, date_val.month, date_val.day)
                        cell = pobs_sheet.cell(row=i, column=data_sped_col_idx, value=only_date)
                        cell.number_format = "DD/MM/YYYY"
                    changed = True

                if changed:
                    updated_rows.append([cell.value for cell in row])
                    updates += 1

        # Format CAP column
        if "CAP" in headers:
            cap_idx = headers.index("CAP") + 1
            for r in range(2, pobs_sheet.max_row + 1):
                cell = pobs_sheet.cell(row=r, column=cap_idx)
                if cell.value:
                    try:
                        cell.value = int(cell.value)
                        cell.number_format = "00000"
                    except:
                        pass

        # Save modifications directly to original POBS
        pobs_wb.save(pobs_path)

        # Also save a copy to POBS CON TRACKING folder
        pobs_tracking_dir = os.path.join(output_dir, "POBS CON TRACKING")
        os.makedirs(pobs_tracking_dir, exist_ok=True)

        # Generate filename for POBS with tracking
        original_name = os.path.splitext(os.path.basename(pobs_path))[0]
        pobs_tracking_filename = f"{original_name}_con_tracking_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        pobs_tracking_path = os.path.join(pobs_tracking_dir, pobs_tracking_filename)
        pobs_wb.save(pobs_tracking_path)

        # Generate TRACKING RADAR file
        if "DATA CONSEGNA" not in headers:
            raise Exception("Column 'DATA CONSEGNA' not found.")
        last_col_idx = headers.index("DATA CONSEGNA") + 1
        radar_headers = headers[:last_col_idx]

        radar_wb = openpyxl.Workbook()
        radar_ws = radar_wb.active
        radar_ws.title = "Tracking Radar"
        radar_ws.append(radar_headers)

        # Process rows before adding to RADAR
        tracking_col_idx_radar = radar_headers.index("TRACKING - LDV TNT") if "TRACKING - LDV TNT" in radar_headers else None

        for row in updated_rows:
            row_data = list(row[:last_col_idx])
            # Clean tracking value if column exists
            if tracking_col_idx_radar is not None:
                tracking_val = row_data[tracking_col_idx_radar]
                if tracking_val:
                    # Remove formula-like formatting
                    cleaned_val = str(tracking_val).replace('="', '').replace('"', '').strip()
                    row_data[tracking_col_idx_radar] = cleaned_val
            radar_ws.append(row_data)

        # Format IMEI* column to ensure NUMBER format
        imei_column_name = None
        for header in radar_headers:
            if "IMEI" in header:
                imei_column_name = header
                break

        if imei_column_name:
            imei_idx = radar_headers.index(imei_column_name) + 1
            for r in range(2, radar_ws.max_row + 1):
                cell = radar_ws.cell(row=r, column=imei_idx)
                if cell.value:
                    try:
                        # Convert to integer to ensure it's stored as number
                        cell.value = int(str(cell.value).strip())
                        # Apply number format with no decimal places
                        cell.number_format = '0'
                    except:
                        pass

        for colname in ["DATA SPEDIZIONE", "DATA CONSEGNA", "Data/ora creazione"]:
            if colname in radar_headers:
                col_idx = radar_headers.index(colname) + 1
                for r in range(2, radar_ws.max_row + 1):
                    cell = radar_ws.cell(row=r, column=col_idx)
                    if isinstance(cell.value, datetime):
                        from datetime import date
                        only_date = date(cell.value.year, cell.value.month, cell.value.day)
                        cell.value = only_date
                        cell.number_format = "DD/MM/YYYY"

        if "CAP" in radar_headers:
            cap_idx = radar_headers.index("CAP") + 1
            for r in range(2, radar_ws.max_row + 1):
                cell = radar_ws.cell(row=r, column=cap_idx)
                if cell.value:
                    try:
                        cell.value = int(cell.value)
                        cell.number_format = "00000"
                    except:
                        pass

        # Create TRACKING RADAR folder
        radar_dir = os.path.join(output_dir, "TRACKING RADAR")
        os.makedirs(radar_dir, exist_ok=True)

        # Generate custom filename if provided
        if custom_name:
            radar_filename = custom_name
        else:
            radar_filename = f"TRACKING RADAR_{datetime.now().strftime('%Y%m%d')}.xlsx"

        radar_output_path = os.path.join(radar_dir, radar_filename)
        radar_wb.save(radar_output_path)

        # Create structured log using new logging system
        log_details = {
            "pobs_file": os.path.basename(pobs_path),
            "transport_file": os.path.basename(trasporti_path),
            "masterfile": os.path.basename(masterfile_path) if masterfile_path else "None",
            "custom_name": custom_name or "Default",
            "total_rows_updated": updates,
            "backup_file": backup_filename,
            "pobs_with_tracking_file": pobs_tracking_filename,
            "tracking_radar_file": radar_filename,
            "tracking_mappings_found": len(mapping_tracking),
            "shipping_dates_found": len(master_dates) if masterfile_path else 0
        }

        files_created = [radar_filename, pobs_tracking_filename, backup_filename]

        log_filename = log_tracking_operation(
            operation_name="UPDATE_TRACKING_DATA",
            status="SUCCESS",
            details=log_details,
            files_created=files_created
        )

        # Keep legacy log for backward compatibility
        log_path = os.path.join(backup_dir, "tracking_updates.log")
        with open(log_path, "a", encoding="utf-8") as f:
            f.write("="*60 + "\n")
            f.write(f"Execution: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"POBS updated directly: {pobs_path}\n")
            f.write(f"POBS with tracking saved: {pobs_tracking_filename}\n")
            f.write(f"Backup saved: {backup_filename}\n")
            f.write(f"TRACKING RADAR file: {radar_filename}\n")
            f.write(f"Total rows updated: {updates}\n\n")

        return {
            'success': True,
            'message': f'Successfully updated {updates} tracking records and generated TRACKING RADAR',
            'updates_count': updates,
            'backup_file': backup_filename,
            'pobs_tracking_file': pobs_tracking_filename,
            'radar_file': radar_filename,
            'radar_path': os.path.join('TRACKING RADAR', radar_filename),
            'download_files': [radar_filename, pobs_tracking_filename],
            'log_file': log_filename
        }

    except Exception as e:
        # Log error
        error_details = {
            "pobs_file": os.path.basename(pobs_path) if pobs_path else "Unknown",
            "transport_file": os.path.basename(trasporti_path) if trasporti_path else "Unknown",
            "masterfile": os.path.basename(masterfile_path) if masterfile_path else "None",
            "custom_name": custom_name or "Default",
            "error_message": str(e),
            "output_directory": output_dir
        }

        log_filename = log_tracking_operation(
            operation_name="UPDATE_TRACKING_DATA",
            status="ERROR",
            details=error_details,
            errors=[str(e)]
        )

        return {
            'success': False,
            'error': str(e),
            'log_file': log_filename
        }

def generate_upload_gsped_realtime(pobs_path, masterfile_path, output_dir, session_id=None):
    """
    Real-time version of generate_upload_gsped with live logging
    """
    if session_id is None:
        session_id = str(uuid.uuid4())

    try:
        realtime_logger.log(session_id, "Starting Upload Gsped generation process...", "info")

        # Create GSPED output directory
        realtime_logger.log(session_id, "Creating GSPED output directory...", "info")
        gsped_dir = os.path.join(output_dir, "GSPED")
        os.makedirs(gsped_dir, exist_ok=True)
        realtime_logger.log(session_id, "GSPED directory ready", "success")

        # Generate progressive filename
        realtime_logger.log(session_id, "Generating output filename...", "info")
        oggi = datetime.now().strftime("%Y%m%d")
        pattern = re.compile(rf"^Upload Gsped_{oggi}_(\d+)\.xls$")

        progressivi = []
        for fname in os.listdir(gsped_dir):
            match = pattern.match(fname)
            if match:
                progressivi.append(int(match.group(1)))
        progressivo = max(progressivi) + 1 if progressivi else 1
        output_filename = f"Upload Gsped_{oggi}_{progressivo:02d}.xls"
        output_path = os.path.join(gsped_dir, output_filename)
        realtime_logger.log(session_id, f"Output filename: {output_filename}", "success")

        # Load masterfile and extract GUIDs from "PER STOPRIPARO" sheet
        realtime_logger.log(session_id, f"Loading master file: {os.path.basename(masterfile_path)}", "info")
        master_wb = openpyxl.load_workbook(masterfile_path, data_only=True)
        if "PER STOPRIPARO" not in master_wb.sheetnames:
            realtime_logger.log(session_id, "MasterFile does not contain 'PER STOPRIPARO' sheet", "error")
            raise Exception("MasterFile does not contain 'PER STOPRIPARO' sheet.")
        master_sheet = master_wb["PER STOPRIPARO"]
        realtime_logger.log(session_id, "Master file loaded successfully", "success")

        realtime_logger.log(session_id, "Extracting GUIDs from master file...", "info")
        master_guids = set()
        for row_idx, row in enumerate(master_sheet.iter_rows(values_only=True), start=1):
            if row_idx == 1:  # Skip header
                continue
            guid = row[1]  # Column B
            if guid:
                master_guids.add(str(guid).strip())
        realtime_logger.log(session_id, f"Extracted {len(master_guids)} GUIDs from master file", "success")

        # Load POBS file
        realtime_logger.log(session_id, f"Loading POBS file: {os.path.basename(pobs_path)}", "info")
        pobs_wb = openpyxl.load_workbook(pobs_path, data_only=True)
        pobs_sheet = pobs_wb.active
        realtime_logger.log(session_id, f"POBS file loaded with {pobs_sheet.max_row} rows", "success")

        # Load template for headers
        realtime_logger.log(session_id, "Loading template headers...", "info")
        try:
            template_wb = xlrd.open_workbook("Gsped_template_excel.xls")
            template_sheet = template_wb.sheet_by_index(0)
            headers = template_sheet.row_values(0)
            realtime_logger.log(session_id, "Template headers loaded from file", "success")
        except FileNotFoundError:
            # Default headers if template not found
            headers = [
                "CD_SOCIETA", "CD_DEPOSITO", "CD_UTENTE", "INDIRIZZO", "CIVICO", "LOCALITA",
                "CAP", "PROVINCIA", "CD_PRODOTTO", "DS_PRODOTTO",
                "QT_ORDINATA", "QT_EVASA", "QT_DA_EVADERE", "QT_SPEDITA", "FG_SPEDIRE",
                "NOTE1", "NOTE2", "RIFERIMENTO_ESTERNO",
                "CAMPO_LIBERO1", "CAMPO_LIBERO2", "CAMPO_LIBERO3", "CAMPO_LIBERO4",
                "CAMPO_LIBERO5", "CAMPO_LIBERO6", "RAGIONE_SOCIALE", "EMAIL", "TELEFONO"
            ]
            realtime_logger.log(session_id, "Using default template headers", "info")

        # Create new XLS workbook
        realtime_logger.log(session_id, "Creating output workbook...", "info")
        new_wb = xlwt.Workbook()
        new_ws = new_wb.add_sheet("Upload Gsped")

        # Write headers
        for col, header in enumerate(headers):
            new_ws.write(0, col, header)
        realtime_logger.log(session_id, f"Headers written ({len(headers)} columns)", "success")

        # Process POBS data
        realtime_logger.log(session_id, "Processing POBS data for Gsped mapping...", "info")
        mapped_rows = []
        r_values = []
        processed_count = 0

        for row_idx, row in enumerate(pobs_sheet.iter_rows(values_only=True), start=1):
            if row_idx == 1:  # Skip header
                continue

            guid = row[7] if len(row) > 7 else None  # Column H (index 7)
            if guid and str(guid).strip() in master_guids:
                if processed_count % 50 == 0 and processed_count > 0:
                    realtime_logger.log(session_id, f"Processed {processed_count} matching records...", "info")
                # Map POBS data to Gsped format (same mapping as original)
                mapped_row = [
                    "392369",           # CD_SOCIETA
                    "",                 # CD_DEPOSITO
                    row[11] if len(row) > 11 else "",  # CD_UTENTE (col L)
                    row[13] if len(row) > 13 else "",  # INDIRIZZO (col N)
                    row[15] if len(row) > 15 else "",  # CIVICO (col P)
                    row[14] if len(row) > 14 else "",  # LOCALITA (col O)
                    row[16] if len(row) > 16 else "",  # CAP (col Q)
                    row[17] if len(row) > 17 else "",  # PROVINCIA (col R)
                    row[20] if len(row) > 20 else "",  # CD_PRODOTTO (col U)
                    row[19] if len(row) > 19 else "",  # DS_PRODOTTO (col T)
                    "2",                # QT_ORDINATA
                    "0",                # QT_EVASA
                    "0",                # QT_DA_EVADERE
                    "1",                # QT_SPEDITA
                    "1",                # FG_SPEDIRE
                    "",                 # NOTE1
                    "",                 # NOTE2
                    row[6] if len(row) > 6 else "",    # RIFERIMENTO_ESTERNO (col G)
                    "",                 # CAMPO_LIBERO1
                    "",                 # CAMPO_LIBERO2
                    "",                 # CAMPO_LIBERO3
                    "",                 # CAMPO_LIBERO4
                    "",                 # CAMPO_LIBERO5
                    "",                 # CAMPO_LIBERO6
                    "VODAFONE EASYRENT", # RAGIONE_SOCIALE
                    "",                 # EMAIL
                    ""                  # TELEFONO
                ]

                # Convert None values to empty strings and strip
                cleaned_row = tuple("" if v is None else str(v).strip() for v in mapped_row)
                mapped_rows.append(cleaned_row)
                r_values.append("" if row[6] is None else str(row[6]).strip())
                processed_count += 1

        realtime_logger.log(session_id, f"Mapped {len(mapped_rows)} records from {processed_count} matching GUIDs", "success")

        # Remove duplicates
        realtime_logger.log(session_id, "Removing duplicate records...", "info")
        unique_rows = list(dict.fromkeys(mapped_rows))
        n_duplicates = len(mapped_rows) - len(unique_rows)
        realtime_logger.log(session_id, f"Removed {n_duplicates} duplicate records, {len(unique_rows)} unique records remaining", "success")

        # Count duplicates by R value (RIFERIMENTO_ESTERNO)
        dup_counter = defaultdict(int)
        seen = set()
        for row, r_val in zip(mapped_rows, r_values):
            if row in seen:
                dup_counter[r_val] += 1
            else:
                seen.add(row)

        # Write unique rows to file
        realtime_logger.log(session_id, "Writing data to output file...", "info")
        row_out = 1
        for mapped_row in unique_rows:
            for col, value in enumerate(mapped_row):
                new_ws.write(row_out, col, value)
            row_out += 1

        new_wb.save(output_path)
        realtime_logger.log(session_id, f"Output file saved: {output_filename}", "success")
        realtime_logger.log(session_id, "Upload Gsped generation completed successfully", "success")

        # Create structured log using new logging system
        log_details = {
            "pobs_file": os.path.basename(pobs_path),
            "masterfile": os.path.basename(masterfile_path),
            "generated_file": output_filename,
            "total_rows_processed": processed_count,
            "total_rows_mapped": len(mapped_rows),
            "unique_rows": len(unique_rows),
            "duplicate_rows_removed": n_duplicates,
            "duplicate_details": dict(dup_counter) if n_duplicates > 0 else {},
            "master_guids_found": len(master_guids),
            "output_directory": gsped_dir
        }

        log_filename = log_tracking_operation(
            operation_name="GENERATE_UPLOAD_GSPED_REALTIME",
            status="SUCCESS",
            details=log_details,
            files_created=[output_filename]
        )

        result = {
            'success': True,
            'message': f'Successfully generated Gsped file with {len(unique_rows)} unique records',
            'output_file': output_filename,
            'processed_count': processed_count,
            'total_rows': len(mapped_rows),
            'unique_rows': len(unique_rows),
            'duplicates_removed': n_duplicates,
            'duplicate_details': dict(dup_counter),
            'download_file': output_filename,
            'log_file': log_filename
        }

        # Store result and complete session
        if session_id:
            realtime_logger.store_result(session_id, result)
            realtime_logger.complete_session(session_id)

        return result

    except Exception as e:
        if session_id:
            realtime_logger.log_error(session_id, f"Operation failed: {str(e)}")

        # Log error
        error_details = {
            "pobs_file": os.path.basename(pobs_path) if pobs_path else "Unknown",
            "masterfile": os.path.basename(masterfile_path) if masterfile_path else "Unknown",
            "error_message": str(e),
            "output_directory": output_dir
        }

        log_filename = log_tracking_operation(
            operation_name="GENERATE_UPLOAD_GSPED_REALTIME",
            status="ERROR",
            details=error_details,
            errors=[str(e)]
        )

        result = {
            'success': False,
            'error': str(e),
            'log_file': log_filename
        }

        # Store error result and complete session
        if session_id:
            realtime_logger.store_result(session_id, result)
            realtime_logger.complete_session(session_id)

        return result

def update_tracking_data_realtime(pobs_path, trasporti_path, masterfile_path, output_dir, custom_name=None, session_id=None):
    """
    Real-time version of update_tracking_data with live logging
    """
    if session_id is None:
        session_id = str(uuid.uuid4())

    try:
        realtime_logger.log(session_id, "Starting tracking data update process...", "info")

        # Create backup first
        realtime_logger.log(session_id, "Creating backup of original POBS file...", "info")
        backup_dir = os.path.join(output_dir, "BACKUP_POBS")
        os.makedirs(backup_dir, exist_ok=True)
        backup_filename = f"{os.path.splitext(os.path.basename(pobs_path))[0]}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        backup_path = os.path.join(backup_dir, backup_filename)
        import shutil
        shutil.copy2(pobs_path, backup_path)
        realtime_logger.log(session_id, f"Backup created: {backup_filename}", "success")

        # Load POBS file
        realtime_logger.log(session_id, f"Loading POBS file: {os.path.basename(pobs_path)}", "info")
        pobs_wb = openpyxl.load_workbook(pobs_path)
        pobs_sheet = pobs_wb.active
        realtime_logger.log(session_id, "POBS file loaded successfully", "success")

        # Load transport file (CSV or Excel)
        realtime_logger.log(session_id, f"Loading transport file: {os.path.basename(trasporti_path)}", "info")
        if trasporti_path.lower().endswith(".csv"):
            trasporti_df = pd.read_csv(trasporti_path, dtype=str, sep=None, engine="python")
        else:
            trasporti_df = pd.read_excel(trasporti_path, dtype=str)

        trasporti_df = trasporti_df.fillna("")
        realtime_logger.log(session_id, f"Transport file loaded with {len(trasporti_df)} rows", "success")

        # Check required columns
        if "Riferimento alfanumerico" not in trasporti_df.columns or "N. sped." not in trasporti_df.columns:
            realtime_logger.log(session_id, "Transport file missing required columns: 'Riferimento alfanumerico' or 'N. sped.'", "error")
            raise Exception("Transport file missing required columns: 'Riferimento alfanumerico' or 'N. sped.'")

        # Create tracking mapping dictionary - clean values from transport file
        realtime_logger.log(session_id, "Processing transport data and creating tracking mappings...", "info")
        def clean_tracking_value(val):
            """Remove formula-like formatting from tracking values"""
            if pd.isna(val):
                return ""
            val_str = str(val).strip()
            # Remove Excel formula formatting like ="value"
            if val_str.startswith('="') and val_str.endswith('"'):
                val_str = val_str[2:-1]
            # Remove any remaining quotes
            val_str = val_str.replace('"', '').strip()
            return val_str

        mapping_tracking = {}
        for _, row in trasporti_df.iterrows():
            ref = str(row["Riferimento alfanumerico"]).strip() if pd.notna(row["Riferimento alfanumerico"]) else ""
            tracking = clean_tracking_value(row["N. sped."])
            if ref:
                mapping_tracking[ref] = tracking

        realtime_logger.log(session_id, f"Created {len(mapping_tracking)} tracking mappings", "success")

        # Load MasterFile for shipping dates
        master_dates = {}
        if masterfile_path:
            realtime_logger.log(session_id, f"Loading master file for shipping dates: {os.path.basename(masterfile_path)}", "info")
            master_wb = openpyxl.load_workbook(masterfile_path, data_only=True)
            if "PER STOPRIPARO" in master_wb.sheetnames:
                sheet = master_wb["PER STOPRIPARO"]
                for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    if i == 1:  # Skip header
                        continue
                    guid = str(row[1]).strip() if row[1] else None
                    data_sped = row[2]  # Column "DATA SPEDIZIONE/BOLLA"
                    if guid:
                        master_dates[guid] = data_sped
                realtime_logger.log(session_id, f"Loaded {len(master_dates)} shipping dates from master file", "success")

        # Find required columns in POBS
        realtime_logger.log(session_id, "Locating required columns in POBS file...", "info")
        headers = [c.value for c in pobs_sheet[1]]
        if "TRACKING - LDV TNT" not in headers or "DATA SPEDIZIONE" not in headers:
            realtime_logger.log(session_id, "Required columns not found in POBS", "error")
            raise Exception("Required columns not found in POBS.")
        tracking_col_idx = headers.index("TRACKING - LDV TNT") + 1
        data_sped_col_idx = headers.index("DATA SPEDIZIONE") + 1
        realtime_logger.log(session_id, "Required columns located successfully", "success")

        updates = 0
        updated_rows = []

        # Update POBS with tracking numbers and shipping dates
        realtime_logger.log(session_id, "Updating POBS with tracking numbers and shipping dates...", "info")
        total_rows = pobs_sheet.max_row - 1  # Exclude header

        for i, row in enumerate(pobs_sheet.iter_rows(min_row=2, values_only=False), start=2):
            if (i - 2) % 100 == 0 and i > 2:
                realtime_logger.log(session_id, f"Processing row {i-1} of {total_rows}...", "info")

            guid_pobs = row[6].value  # Column G
            if guid_pobs:
                guid_str = str(guid_pobs).strip()
                changed = False

                # Update tracking - clean value to avoid formula-like formatting
                if guid_str in mapping_tracking:
                    tracking_value = mapping_tracking[guid_str].strip()
                    # Remove any quotes or special characters that might cause Excel formula issues
                    tracking_value = tracking_value.replace('="', '').replace('"', '').strip()
                    pobs_sheet.cell(row=i, column=tracking_col_idx, value=tracking_value)
                    changed = True

                # Update shipping date from masterfile
                if guid_str in master_dates:
                    date_val = master_dates[guid_str]
                    if isinstance(date_val, datetime):
                        from datetime import date
                        only_date = date(date_val.year, date_val.month, date_val.day)
                        cell = pobs_sheet.cell(row=i, column=data_sped_col_idx, value=only_date)
                        cell.number_format = "DD/MM/YYYY"
                    changed = True

                if changed:
                    updated_rows.append([cell.value for cell in row])
                    updates += 1

        realtime_logger.log(session_id, f"Updated {updates} records in POBS", "success")

        # Format CAP column
        realtime_logger.log(session_id, "Formatting CAP column...", "info")
        if "CAP" in headers:
            cap_idx = headers.index("CAP") + 1
            for r in range(2, pobs_sheet.max_row + 1):
                cell = pobs_sheet.cell(row=r, column=cap_idx)
                if cell.value:
                    try:
                        cell.value = int(cell.value)
                        cell.number_format = "00000"
                    except:
                        pass

        # Save modifications directly to original POBS
        realtime_logger.log(session_id, "Saving updated POBS file...", "info")
        pobs_wb.save(pobs_path)

        # Also save a copy to POBS CON TRACKING folder
        pobs_tracking_dir = os.path.join(output_dir, "POBS CON TRACKING")
        os.makedirs(pobs_tracking_dir, exist_ok=True)

        # Generate filename for POBS with tracking
        original_name = os.path.splitext(os.path.basename(pobs_path))[0]
        pobs_tracking_filename = f"{original_name}_con_tracking_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        pobs_tracking_path = os.path.join(pobs_tracking_dir, pobs_tracking_filename)
        pobs_wb.save(pobs_tracking_path)
        realtime_logger.log(session_id, f"POBS with tracking saved: {pobs_tracking_filename}", "success")

        # Generate TRACKING RADAR file
        realtime_logger.log(session_id, "Generating TRACKING RADAR file...", "info")
        if "DATA CONSEGNA" not in headers:
            realtime_logger.log(session_id, "Column 'DATA CONSEGNA' not found", "error")
            raise Exception("Column 'DATA CONSEGNA' not found.")
        last_col_idx = headers.index("DATA CONSEGNA") + 1
        radar_headers = headers[:last_col_idx]

        radar_wb = openpyxl.Workbook()
        radar_ws = radar_wb.active
        radar_ws.title = "Tracking Radar"
        radar_ws.append(radar_headers)

        # Process rows before adding to RADAR
        tracking_col_idx_radar = radar_headers.index("TRACKING - LDV TNT") if "TRACKING - LDV TNT" in radar_headers else None

        for row in updated_rows:
            row_data = list(row[:last_col_idx])
            # Clean tracking value if column exists
            if tracking_col_idx_radar is not None:
                tracking_val = row_data[tracking_col_idx_radar]
                if tracking_val:
                    # Remove formula-like formatting
                    cleaned_val = str(tracking_val).replace('="', '').replace('"', '').strip()
                    row_data[tracking_col_idx_radar] = cleaned_val
            radar_ws.append(row_data)

        # Format IMEI* column to ensure NUMBER format
        imei_column_name = None
        for header in radar_headers:
            if "IMEI" in header:
                imei_column_name = header
                break

        if imei_column_name:
            imei_idx = radar_headers.index(imei_column_name) + 1
            for r in range(2, radar_ws.max_row + 1):
                cell = radar_ws.cell(row=r, column=imei_idx)
                if cell.value:
                    try:
                        # Convert to integer to ensure it's stored as number
                        cell.value = int(str(cell.value).strip())
                        # Apply number format with no decimal places
                        cell.number_format = '0'
                    except:
                        pass

        for colname in ["DATA SPEDIZIONE", "DATA CONSEGNA", "Data/ora creazione"]:
            if colname in radar_headers:
                col_idx = radar_headers.index(colname) + 1
                for r in range(2, radar_ws.max_row + 1):
                    cell = radar_ws.cell(row=r, column=col_idx)
                    if isinstance(cell.value, datetime):
                        from datetime import date
                        only_date = date(cell.value.year, cell.value.month, cell.value.day)
                        cell.value = only_date
                        cell.number_format = "DD/MM/YYYY"

        if "CAP" in radar_headers:
            cap_idx = radar_headers.index("CAP") + 1
            for r in range(2, radar_ws.max_row + 1):
                cell = radar_ws.cell(row=r, column=cap_idx)
                if cell.value:
                    try:
                        cell.value = int(cell.value)
                        cell.number_format = "00000"
                    except:
                        pass

        # Create TRACKING RADAR folder
        radar_dir = os.path.join(output_dir, "TRACKING RADAR")
        os.makedirs(radar_dir, exist_ok=True)

        # Generate custom filename if provided
        if custom_name:
            radar_filename = custom_name
        else:
            radar_filename = f"TRACKING RADAR_{datetime.now().strftime('%Y%m%d')}.xlsx"

        radar_output_path = os.path.join(radar_dir, radar_filename)
        radar_wb.save(radar_output_path)
        realtime_logger.log(session_id, f"TRACKING RADAR saved: {radar_filename}", "success")

        # Create structured log using new logging system
        log_details = {
            "pobs_file": os.path.basename(pobs_path),
            "transport_file": os.path.basename(trasporti_path),
            "masterfile": os.path.basename(masterfile_path) if masterfile_path else "None",
            "custom_name": custom_name or "Default",
            "total_rows_updated": updates,
            "backup_file": backup_filename,
            "pobs_with_tracking_file": pobs_tracking_filename,
            "tracking_radar_file": radar_filename,
            "tracking_mappings_found": len(mapping_tracking),
            "shipping_dates_found": len(master_dates) if masterfile_path else 0
        }

        files_created = [radar_filename, pobs_tracking_filename, backup_filename]

        log_filename = log_tracking_operation(
            operation_name="UPDATE_TRACKING_DATA_REALTIME",
            status="SUCCESS",
            details=log_details,
            files_created=files_created
        )

        realtime_logger.log(session_id, "Tracking data update completed successfully", "success")

        result = {
            'success': True,
            'message': f'Successfully updated {updates} tracking records and generated TRACKING RADAR',
            'updates_count': updates,
            'backup_file': backup_filename,
            'pobs_tracking_file': pobs_tracking_filename,
            'radar_file': radar_filename,
            'radar_path': os.path.join('TRACKING RADAR', radar_filename),
            'download_files': [radar_filename, pobs_tracking_filename],
            'log_file': log_filename
        }

        # Store error result and complete session
        if session_id:
            realtime_logger.store_result(session_id, result)
            realtime_logger.complete_session(session_id)

        return result

    except Exception as e:
        if session_id:
            realtime_logger.log_error(session_id, f"Operation failed: {str(e)}")

        # Log error
        error_details = {
            "pobs_file": os.path.basename(pobs_path) if pobs_path else "Unknown",
            "transport_file": os.path.basename(trasporti_path) if trasporti_path else "Unknown",
            "masterfile": os.path.basename(masterfile_path) if masterfile_path else "None",
            "custom_name": custom_name or "Default",
            "error_message": str(e),
            "output_directory": output_dir
        }

        log_filename = log_tracking_operation(
            operation_name="UPDATE_TRACKING_DATA_REALTIME",
            status="ERROR",
            details=error_details,
            errors=[str(e)]
        )

        result = {
            'success': False,
            'error': str(e),
            'log_file': log_filename
        }

        # Store error result and complete session
        if session_id:
            realtime_logger.store_result(session_id, result)
            realtime_logger.complete_session(session_id)

        return result