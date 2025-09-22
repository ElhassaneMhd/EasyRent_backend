"""
PCOM Service - Converted from EasyRent_PCOM_App.py
Removes Tkinter GUI and converts to web service functions
"""

import os
import re
import shutil
import uuid
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from .logger_service import log_pcom_operation
from .realtime_logger import realtime_logger

def extract_memory(text: str):
    """Extract memory information from text"""
    if not isinstance(text, str):
        return None
    m = re.search(r"(\d+)\s*GB", text, re.IGNORECASE)
    if m:
        return m.group(1) + "GB"
    return None

def clean_model(text: str):
    """Clean model text"""
    if not isinstance(text, str):
        return None
    text = re.sub(r"\+.*", "", text)  # remove "+ accessories"
    text = re.sub(r"\s*\d+\s*GB", "", text, flags=re.IGNORECASE)  # remove memory
    return text.strip()

def load_mapping(modelli_path=None):
    """Load model mapping from Excel file"""
    if modelli_path and os.path.isfile(modelli_path):
        df = pd.read_excel(modelli_path, sheet_name=0)
    else:
        raise FileNotFoundError("File Modelli Easyrent.xlsx missing")

    mapping = pd.DataFrame()
    mapping["Versione"] = df["Edition"]
    mapping["Modello"] = df["Modello"].apply(clean_model)
    mapping["Memoria"] = df["Modello"].apply(extract_memory)
    mapping["Accessori"] = df["Edition"].apply(
        lambda x: "Alimentatore e auricolari originali" if isinstance(x, str) and "KE" in str(x) else ""
    )
    return mapping.drop_duplicates().set_index("Versione").to_dict(orient="index")

def process_pcom_with_pobs(noleggio_path, soho_path, pobs_path, output_dir, modelli_path, options, custom_names=None):
    """
    Process PCOM files and optionally update POBS
    Includes new features from updated script 1
    """
    try:
        results = {
            'success': True,
            'message': 'Processing completed'
        }

        # Process PCOM
        pcom_result = process_pcom_files(noleggio_path, soho_path, output_dir, modelli_path, options, custom_names.get('pcom') if custom_names else None)
        if not pcom_result.get('success', True):
            return {
                'success': False,
                'error': f"PCOM processing failed: {pcom_result.get('error', 'Unknown error')}"
            }
        results['pcom'] = pcom_result

        # Process POBS if path provided
        if pobs_path:
            pobs_result = process_pobs_update(pobs_path, noleggio_path, output_dir, custom_names.get('pobs') if custom_names else None)
            if not pobs_result.get('success', True):
                return {
                    'success': False,
                    'error': f"POBS processing failed: {pobs_result.get('error', 'Unknown error')}"
                }
            results['pobs'] = pobs_result
            results['message'] = f"PCOM and POBS processing completed. PCOM: {pcom_result.get('records_processed', 0)} records, POBS: {pobs_result.get('records_added', 0)} records added"
        else:
            results['message'] = f"PCOM processing completed. {pcom_result.get('records_processed', 0)} records processed"
            results['records_processed'] = pcom_result.get('records_processed', 0)
            results['output_file'] = pcom_result.get('output_file')
            results['download_file'] = pcom_result.get('download_file')
            results['processing_log'] = pcom_result.get('processing_log', [])

        return results
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

def process_pobs_update(pobs_path, noleggio_path, dest_folder, custom_name):
    """
    Update POBS file with new records from Noleggio
    New function from updated script 1
    """
    try:
        processing_log = []

        def log(message):
            processing_log.append(message)

        log("[POBS] Opening POBS history file...")
        wb_pobs = load_workbook(pobs_path)
        ws_pobs = wb_pobs.active
        headers = [c.value for c in ws_pobs[1]]

        log("[POBS] Opening Noleggio file...")
        wb_noleggio = load_workbook(noleggio_path, data_only=True)
        ws_nol = wb_noleggio.active
        headers_nol = [c.value for c in ws_nol[1]]

        # Column mapping from script 1
        mapping = {
            "ID Soluzione Digitale": "ID Soluzione Digitale",
            "Nome Partner Logistica": "Nome Partner Logistica",
            "Soluzione Digitale": "Soluzione Digitale",
            "ID Versione": "ID Versione",
            "Versione": "Versione",
            "ID Opportunità/ Pratica BSales": "ID Opportunità/ Pratica BSales",
            "POBS ID": "POBS ID",
            "GUID": "GUID",
            "Codice Cliente": "Codice Cliente",
            "IMEI*": "IMEI*",
            "Data/ora creazione": "Data assegnazione",
            "Ragione Sociale": "ragione Sociale",
            "Partita IVA": "P.IVA",
            "Full Address": "indirizzo",
            "Città": "Citta'",
            "Codice Postale": "CAP",
            "provincia": "provincia",
            "Nome referente": "Nome referente PdA",
            "Cognome referente": "cognome referente PdA",
            "E-mail cliente": "E-mail",
            "Telefono cliente": "Numero cellulare",
            "TRACKING - LDV TNT": None,
            "DATA SPEDIZIONE": None,
            "DATA CONSEGNA": None,
            "STATO": "IN GESTIONE",
            "DATA RIENTRO": None,
            "CAUSALE GIACENZA/RIENTRO": None,
            "NOTE": None,
            "TRACKING DISATTIVAZIONE": None,
            "DATA RIENTRO.1": None,
        }

        log("[POBS] Adding new rows...")
        records_added = 0
        for r in range(2, ws_nol.max_row+1):
            new_row = []
            for col_pobs, col_nol in mapping.items():
                if col_nol is None:
                    val = ""
                elif col_nol == "IN GESTIONE":
                    val = "IN GESTIONE"
                else:
                    col_idx = headers_nol.index(col_nol)+1 if col_nol in headers_nol else None
                    val = ws_nol.cell(row=r, column=col_idx).value if col_idx else ""
                    # Format date if needed
                    if col_pobs == "Data/ora creazione" and val:
                        if hasattr(val, "strftime"):
                            val = val.strftime("%d/%m/%Y")
                new_row.append(val)
            ws_pobs.append(new_row)
            records_added += 1

        # Create backup
        backup_dir = Path(dest_folder) / "backup_POBS"
        backup_dir.mkdir(parents=True, exist_ok=True)
        shutil.copy(str(pobs_path), str(backup_dir / Path(pobs_path).name))

        # Create POBS output directory
        pobs_dir = os.path.join(dest_folder, "POBS")
        os.makedirs(pobs_dir, exist_ok=True)

        # Generate output filename
        today = datetime.now().strftime("%Y%m%d")
        default_name = f"POBS_aggiornato_{today}.xlsx"
        out_name = custom_name if custom_name else default_name
        out_path = os.path.join(pobs_dir, out_name)

        log(f"[POBS] Saving to: {out_path}")
        wb_pobs.save(out_path)

        return {
            'success': True,
            'message': f'Successfully added {records_added} records to POBS',
            'records_added': records_added,
            'output_file': out_name,
            'processing_log': processing_log,
            'download_file': out_name
        }

    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

def process_pcom_files(noleggio_path, soho_path, output_dir, modelli_path, options, custom_name=None):
    """
    Process PCOM files
    Updated from original with custom naming support
    """
    try:
        processing_log = []

        def log(message):
            processing_log.append(message)

        log("[INFO] Starting PCOM processing...")

        # Load mapping if modelli file is provided and modelli option is enabled
        map_dict = {}
        if options.get("modelli", False) and modelli_path:
            log("[INFO] Loading model mapping...")
            try:
                map_dict = load_mapping(modelli_path)
                log(f"[INFO] Loaded {len(map_dict)} model mappings")
            except Exception as e:
                log(f"[WARNING] Could not load model mapping: {str(e)}")

        log("[INFO] Opening SOHO file...")
        # Open SOHO file and find the correct sheet
        wb_soho = load_workbook(soho_path, data_only=True)
        soho_sheet = None
        for name in wb_soho.sheetnames:
            if "Modulo Ordini" in name:
                soho_sheet = name
                break
        if not soho_sheet:
            soho_sheet = wb_soho.sheetnames[0]
            log(f"[INFO] Using first sheet: {soho_sheet}")
        else:
            log(f"[INFO] Using sheet: {soho_sheet}")

        ws_soho = wb_soho[soho_sheet]

        # Build SOHO mappings
        soho_map_notes = {}
        soho_map_imei = {}
        for row in range(10, ws_soho.max_row+1):
            id_val = ws_soho.cell(row=row, column=1).value   # col A
            note_val = ws_soho.cell(row=row, column=8).value # col H
            imei_val = ws_soho.cell(row=row, column=9).value # col I
            if id_val:
                pid = str(id_val).strip()
                if note_val not in (None, ""):
                    soho_map_notes[pid] = note_val
                if imei_val not in (None, ""):
                    soho_map_imei[pid] = str(imei_val).strip()

        log(f"[INFO] Loaded {len(soho_map_notes)} notes and {len(soho_map_imei)} IMEI mappings from SOHO")

        log("[INFO] Opening Noleggio file...")
        wb_noleggio = load_workbook(noleggio_path)
        ws = wb_noleggio.active

        last_col = ws.max_column
        records_processed = 0

        # Add new columns if needed
        if options.get("modelli", False) or options.get("rientro", False):
            ws.cell(row=1, column=last_col+1, value="Modello")
            ws.cell(row=1, column=last_col+2, value="Memoria")
            ws.cell(row=1, column=last_col+3, value="Accessori")
            ws.cell(row=1, column=last_col+4, value="Rientro da Pobs")

        # Process each row
        for row in range(2, ws.max_row+1):
            versione = ws.cell(row=row, column=5).value  # col E
            pobs_id = ws.cell(row=row, column=7).value   # col G
            modello, memoria, accessori, rientro = "", "", "", ""

            # Handle model mapping
            if options.get("modelli", False) and versione in map_dict:
                modello = map_dict[versione]["Modello"] or ""
                memoria = map_dict[versione]["Memoria"] or ""
                accessori = map_dict[versione]["Accessori"] or ""

            # Handle POBS ID related operations
            if pobs_id:
                pid = str(pobs_id).strip()
                if options.get("rientro", False) and pid in soho_map_notes:
                    rientro = soho_map_notes[pid]
                if options.get("imei", False) and pid in soho_map_imei:
                    ws.cell(row=row, column=10, value=soho_map_imei[pid])  # col J = IMEI

            # Add new column data
            if options.get("modelli", False) or options.get("rientro", False):
                ws.cell(row=row, column=last_col+1, value=modello)
                ws.cell(row=row, column=last_col+2, value=memoria)
                ws.cell(row=row, column=last_col+3, value=accessori)
                ws.cell(row=row, column=last_col+4, value=rientro)

            records_processed += 1

        log(f"[INFO] Processed {records_processed} records")

        # Check for empty cells in critical columns and warn user
        empty_cell_warnings = []
        empty_cell_count = 0
        critical_columns = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]  # A through J

        log("[INFO] Checking for empty cells in critical columns...")
        for row in range(2, ws.max_row + 1):
            for col in critical_columns:
                cell_value = ws.cell(row=row, column=col).value
                if cell_value is None or cell_value == '' or (isinstance(cell_value, str) and cell_value.strip() == ''):
                    empty_cell_count += 1

        if empty_cell_count > 0:
            warning_msg = f"⚠️ Warning: Found {empty_cell_count} empty cells in critical columns (A-J). Please review the output file for missing data."
            log(f"[WARNING] {warning_msg}")
            empty_cell_warnings.append(warning_msg)

        # Delete unnecessary columns if requested
        if options.get("clean", False):
            log("[INFO] Cleaning unnecessary columns...")
            from openpyxl.utils import column_index_from_string
            cols_to_delete = ["AB","AA","Z","Y","X","W","V","U","T","M","L","K","I","F","E","D","C","B","A"]
            deleted_count = 0
            for col in cols_to_delete:
                try:
                    idx = column_index_from_string(col)
                    if idx <= ws.max_column:
                        ws.delete_cols(idx, 1)
                        deleted_count += 1
                except:
                    pass
            log(f"[INFO] Deleted {deleted_count} unnecessary columns")

        # Create PCOM output directory
        pcom_dir = os.path.join(output_dir, "PCOM")
        os.makedirs(pcom_dir, exist_ok=True)

        # Generate output filename
        if custom_name:
            output_filename = custom_name
        else:
            base_name = Path(noleggio_path).stem
            parts = base_name.split("_")
            timestamp = parts[-2] + "_" + parts[-1] if len(parts) >= 2 else base_name
            output_filename = f"Ordine_EasyRent_PCOM_{timestamp}.xlsx"
        output_path = os.path.join(pcom_dir, output_filename)

        log(f"[INFO] Saving to: {output_path}")
        wb_noleggio.save(output_path)

        log("[INFO] PCOM processing completed successfully")

        result_message = f'Successfully processed {records_processed} records'
        if empty_cell_warnings:
            result_message += f'. {empty_cell_warnings[0]}'

        return {
            'success': True,
            'message': result_message,
            'records_processed': records_processed,
            'output_file': output_filename,
            'processing_log': processing_log,
            'download_file': output_filename,
            'warnings': empty_cell_warnings
        }

    except Exception as e:
        return {
            'success': False,
            'error': str(e),
            'processing_log': processing_log
        }

def process_pcom_with_pobs_realtime(noleggio_path, soho_path, pobs_path, output_dir, modelli_path, options, custom_names=None, session_id=None):
    """
    Process PCOM files with real-time logging and optionally update POBS
    """
    processing_log = []

    def log_message(message):
        processing_log.append(message)
        if session_id:
            if message.startswith('[INFO]'):
                realtime_logger.log_info(session_id, message[6:])
            elif message.startswith('[OK]'):
                realtime_logger.log_ok(session_id, message[4:])
            elif message.startswith('[WARNING]'):
                realtime_logger.log_warning(session_id, message[9:])
            elif message.startswith('[ERROR]'):
                realtime_logger.log_error(session_id, message[7:])
            else:
                realtime_logger.log_info(session_id, message)

    try:
        log_message("[INFO] Starting PCOM processing with real-time logging...")

        results = {
            'success': True,
            'message': 'Processing completed'
        }

        # Process PCOM
        log_message("[INFO] Processing PCOM files...")
        pcom_result = process_pcom_files_realtime(noleggio_path, soho_path, output_dir, modelli_path, options, custom_names.get('pcom') if custom_names else None, session_id)
        if not pcom_result.get('success', True):
            log_message(f"[ERROR] PCOM processing failed: {pcom_result.get('error', 'Unknown error')}")
            if session_id:
                realtime_logger.complete_session(session_id)
            return {
                'success': False,
                'error': f"PCOM processing failed: {pcom_result.get('error', 'Unknown error')}",
                'processing_log': processing_log
            }
        results['pcom'] = pcom_result
        log_message(f"[OK] PCOM processing completed: {pcom_result.get('records_processed', 0)} records processed")

        # Process POBS if path provided
        if pobs_path:
            log_message("[INFO] Processing POBS update...")
            pobs_result = process_pobs_update_realtime(pobs_path, noleggio_path, output_dir, custom_names.get('pobs') if custom_names else None, session_id)
            if not pobs_result.get('success', True):
                log_message(f"[ERROR] POBS processing failed: {pobs_result.get('error', 'Unknown error')}")
                if session_id:
                    realtime_logger.complete_session(session_id)
                return {
                    'success': False,
                    'error': f"POBS processing failed: {pobs_result.get('error', 'Unknown error')}",
                    'processing_log': processing_log
                }
            results['pobs'] = pobs_result
            log_message(f"[OK] POBS processing completed: {pobs_result.get('records_added', 0)} records added")
            results['message'] = f"PCOM and POBS processing completed. PCOM: {pcom_result.get('records_processed', 0)} records, POBS: {pobs_result.get('records_added', 0)} records added"
        else:
            results['message'] = f"PCOM processing completed. {pcom_result.get('records_processed', 0)} records processed"
            results['records_processed'] = pcom_result.get('records_processed', 0)
            results['output_file'] = pcom_result.get('output_file')
            results['download_file'] = pcom_result.get('download_file')
            results['processing_log'] = processing_log

        log_message("[OK] All processing completed successfully")

        if session_id:
            realtime_logger.store_result(session_id, results)
            realtime_logger.complete_session(session_id)

        return results

    except Exception as e:
        log_message(f"[ERROR] Operation failed: {str(e)}")
        if session_id:
            realtime_logger.complete_session(session_id)
        return {
            'success': False,
            'error': str(e),
            'processing_log': processing_log
        }

def process_pcom_files_realtime(noleggio_path, soho_path, output_dir, modelli_path, options, custom_name=None, session_id=None):
    """
    Process PCOM files with real-time logging
    """
    processing_log = []

    def log_message(message):
        processing_log.append(message)
        if session_id:
            if message.startswith('[INFO]'):
                realtime_logger.log_info(session_id, message[6:])
            elif message.startswith('[OK]'):
                realtime_logger.log_ok(session_id, message[4:])
            elif message.startswith('[WARNING]'):
                realtime_logger.log_warning(session_id, message[9:])
            elif message.startswith('[ERROR]'):
                realtime_logger.log_error(session_id, message[7:])
            else:
                realtime_logger.log_info(session_id, message)

    try:
        log_message("[INFO] Starting PCOM file processing...")

        # Ensure output directory exists
        pcom_dir = os.path.join(output_dir, 'PCOM')
        os.makedirs(pcom_dir, exist_ok=True)

        # Load files
        log_message(f"Loading Noleggio file: {os.path.basename(noleggio_path)}")
        noleggio_data = pd.read_excel(noleggio_path)
        log_message(f"Loaded {len(noleggio_data)} records from Noleggio file")

        log_message(f"Loading SOHO file: {os.path.basename(soho_path)}")
        soho_data = pd.read_excel(soho_path)
        log_message(f"Loaded {len(soho_data)} records from SOHO file")

        # Load mapping if provided
        mapping = {}
        if modelli_path and options.get('modelli', True):
            log_message(f"Loading models mapping: {os.path.basename(modelli_path)}")
            mapping = load_mapping(modelli_path)
            log_message(f"Loaded {len(mapping)} model mappings")

        # Load workbook to preserve formatting
        log_message("Processing data and applying transformations...")
        wb_noleggio = load_workbook(noleggio_path)
        ws_noleggio = wb_noleggio.active

        records_processed = 0

        # Process each row (skip header)
        for row in range(2, len(noleggio_data) + 2):
            # Process IMEI
            if options.get('imei', True):
                imei_cell = ws_noleggio[f'J{row}']
                if imei_cell.value:
                    imei_clean = str(imei_cell.value).strip()
                    ws_noleggio[f'J{row}'] = imei_clean

            # Process model mapping
            if options.get('modelli', True) and mapping:
                model_cell = ws_noleggio[f'I{row}']
                if model_cell.value and str(model_cell.value) in mapping:
                    ws_noleggio[f'I{row}'] = mapping[str(model_cell.value)]

            records_processed += 1

            # Log progress every 100 records
            if records_processed % 100 == 0:
                log_message(f"Processed {records_processed} records...")

        log_message(f"Completed processing {records_processed} records")

        # Check for empty cells in critical columns and warn user
        empty_cell_warnings = []
        empty_cell_count = 0
        critical_columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']

        log_message("Checking for empty cells in critical columns...")
        for row in range(2, len(noleggio_data) + 2):
            for col in critical_columns:
                cell_value = ws_noleggio[f'{col}{row}'].value
                if cell_value is None or cell_value == '' or (isinstance(cell_value, str) and cell_value.strip() == ''):
                    empty_cell_count += 1

        if empty_cell_count > 0:
            warning_msg = f"⚠️ Warning: Found {empty_cell_count} empty cells in critical columns (A-J). Please review the output file for missing data."
            log_message(f"[WARNING] {warning_msg}")
            empty_cell_warnings.append(warning_msg)

        # Generate output filename
        if custom_name:
            output_filename = custom_name
        else:
            base_name = Path(noleggio_path).stem
            parts = base_name.split("_")
            timestamp = parts[-2] + "_" + parts[-1] if len(parts) >= 2 else base_name
            output_filename = f"Ordine_EasyRent_PCOM_{timestamp}.xlsx"
        output_path = os.path.join(pcom_dir, output_filename)

        log_message(f"Saving to: {output_filename}")
        wb_noleggio.save(output_path)
        log_message("PCOM file saved successfully")

        result_message = f'Successfully processed {records_processed} records'
        if empty_cell_warnings:
            result_message += f'. {empty_cell_warnings[0]}'

        result = {
            'success': True,
            'message': result_message,
            'records_processed': records_processed,
            'output_file': output_filename,
            'processing_log': processing_log,
            'download_file': output_filename,
            'warnings': empty_cell_warnings
        }

        # Store result and complete session for real-time operations
        if session_id:
            realtime_logger.store_result(session_id, result)
            realtime_logger.complete_session(session_id)

        return result

    except Exception as e:
        log_message(f"Processing failed: {str(e)}")
        result = {
            'success': False,
            'error': str(e),
            'processing_log': processing_log
        }

        # Store error result and complete session for real-time operations
        if session_id:
            realtime_logger.store_result(session_id, result)
            realtime_logger.complete_session(session_id)

        return result

def process_pobs_update_realtime(pobs_path, noleggio_path, dest_folder, custom_name, session_id=None):
    """
    Update POBS file with new records from Noleggio with real-time logging
    """
    processing_log = []

    def log_message(message):
        processing_log.append(message)
        if session_id:
            if message.startswith('[INFO]'):
                realtime_logger.log_info(session_id, message[6:])
            elif message.startswith('[OK]'):
                realtime_logger.log_ok(session_id, message[4:])
            elif message.startswith('[WARNING]'):
                realtime_logger.log_warning(session_id, message[9:])
            elif message.startswith('[ERROR]'):
                realtime_logger.log_error(session_id, message[7:])
            else:
                realtime_logger.log_info(session_id, message)

    try:
        log_message("[INFO] Starting POBS update process...")

        # Read files
        log_message(f"[INFO] Reading POBS file: {os.path.basename(pobs_path)}")
        df_pobs = pd.read_excel(pobs_path, dtype=str)
        original_count = len(df_pobs)
        log_message(f"[OK] Loaded {original_count} records from POBS file")

        log_message(f"[INFO] Reading Noleggio file: {os.path.basename(noleggio_path)}")
        df_noleggio = pd.read_excel(noleggio_path, dtype=str)
        log_message(f"[OK] Loaded {len(df_noleggio)} records from Noleggio file")

        # Find new records
        chiave = "POBS ID"
        log_message("[INFO] Finding new records to add...")

        if chiave not in df_noleggio.columns:
            log_message(f"[ERROR] Column '{chiave}' not found in Noleggio file")
            return {
                'success': False,
                'error': f"Column '{chiave}' not found in Noleggio file",
                'processing_log': processing_log
            }

        if chiave not in df_pobs.columns:
            log_message(f"[ERROR] Column '{chiave}' not found in POBS file")
            return {
                'success': False,
                'error': f"Column '{chiave}' not found in POBS file",
                'processing_log': processing_log
            }

        # Clean data
        df_noleggio[chiave] = df_noleggio[chiave].astype(str).str.strip().str.upper()
        df_pobs[chiave] = df_pobs[chiave].astype(str).str.strip().str.upper()

        # Find new records
        nuovi = df_noleggio[~df_noleggio[chiave].isin(df_pobs[chiave])]

        if nuovi.empty:
            log_message("[INFO] No new records found to add")
            return {
                'success': True,
                'no_changes': True,
                'message': 'No new records to add',
                'records_added': 0,
                'processing_log': processing_log
            }

        log_message(f"[OK] Found {len(nuovi)} new records to add")

        # Append new records
        df_combined = pd.concat([df_pobs, nuovi], ignore_index=True)
        final_count = len(df_combined)
        records_added = len(nuovi)

        # Generate output filename
        if custom_name:
            output_filename = custom_name
        else:
            output_filename = f"POBS_updated_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        output_path = os.path.join(dest_folder, output_filename)

        log_message(f"[INFO] Saving updated POBS file: {output_filename}")
        df_combined.to_excel(output_path, index=False)

        log_message(f"[OK] Successfully added {records_added} records")
        log_message(f"[OK] Final POBS file has {final_count} records")

        return {
            'success': True,
            'message': f'Successfully added {records_added} new records',
            'records_added': records_added,
            'original_count': original_count,
            'final_count': final_count,
            'download_file': output_filename,
            'processing_log': processing_log
        }

    except Exception as e:
        log_message(f"[ERROR] POBS update failed: {str(e)}")
        return {
            'success': False,
            'error': str(e),
            'processing_log': processing_log
        }